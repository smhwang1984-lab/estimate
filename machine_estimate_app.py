import json
import os
import re
import subprocess
import sys
import tkinter as tk
from copy import copy
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

import openpyxl

APP_VERSION = "0.0.4"


class MachineEstimateApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"기계 시트 표준 견적 입력 시스템 v{APP_VERSION}")
        self.root.geometry("1380x760")
        self.root.minsize(1180, 680)
        self.data = []
        self.display_limit = 40
        self.display_page_size = 40
        self.session_restored_count = 0
        self.session_saved_at = None
        self.is_saving = False
        self.search_after_id = None
        self.selected_nos = set()
        # TSERP(py/web/style.css)와 동일 계열의 딥 차콜/슬레이트 팔레트.
        self.colors = {
            "bg": "#11161c", "panel": "#171d25", "panel_2": "#234060", "card": "#18212b",
            "card_alt": "#1a2535", "line": "#2a3340", "text": "#dde4ec", "muted": "#8b97a7",
            "accent": "#4fb0ff", "accent_2": "#9cc8ff", "success_bg": "#2c5c44",
            "success_fg": "#7ddc9e", "warn_bg": "#5a3a1a", "warn_fg": "#ffb648",
            "danger_bg": "#40222a", "danger_fg": "#ff9aa2",
            "new_bg": "#5a3a1a", "new_fg": "#ffb648", "new_border": "#8a5a2a",
        }
        self.rates = {
            "m_5axis": 70000, "m_4axis": 50000, "m_3axis": 40000, "m_lathe": 35000,
            "m_general": 20000, "m_finish": 15600, "m_cmm": 30000, "m_grind": 35000,
            "m_jig": 35000, "m_prog": 35000,
        }
        self.setup_ui_scale()
        self.maximize_window()
        self.build_dashboard()
        self.restore_session_state()
        self.refresh_table(True)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.after(800, self.check_for_update)

    def on_close(self):
        self.save_session_state()
        self.root.destroy()

    def maximize_window(self):
        try:
            self.root.state("zoomed")
        except tk.TclError:
            pass

    def setup_ui_scale(self):
        self.font_family = "맑은 고딕"
        self.font_normal = (self.font_family, 13)
        self.font_bold = (self.font_family, 13, "bold")
        self.font_small = (self.font_family, 11, "bold")
        self.root.option_add("*Font", self.font_normal)
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        c = self.colors
        # 팝업창도 메인 화면과 같은 TSERP 딥블루로 보이도록 ttk 기본 스타일 전체를 맞춘다.
        style.configure("TFrame", background=c["panel"])
        style.configure("TLabel", background=c["panel"], foreground=c["text"])
        style.configure("TButton", padding=(10, 6), background=c["panel_2"], foreground=c["text"], bordercolor=c["line"], focuscolor=c["accent"])
        style.map("TButton", background=[("active", c["accent"]), ("pressed", c["accent"])], foreground=[("active", c["bg"]), ("pressed", c["bg"])])
        style.configure("TEntry", padding=5, fieldbackground=c["card_alt"], foreground=c["text"], insertcolor=c["text"], bordercolor=c["line"])
        style.configure("TCombobox", padding=5, fieldbackground=c["card_alt"], background=c["card_alt"], foreground=c["text"], arrowcolor=c["accent_2"], bordercolor=c["line"])
        style.map("TCombobox", fieldbackground=[("readonly", c["card_alt"])], foreground=[("readonly", c["text"])], background=[("readonly", c["card_alt"])])
        style.configure("TRadiobutton", background=c["panel"], foreground=c["text"])
        style.map("TRadiobutton", background=[("active", c["panel"])], foreground=[("active", c["accent_2"])])
        style.configure("TCheckbutton", background=c["panel"], foreground=c["text"])
        style.map("TCheckbutton", background=[("active", c["panel"])])
        style.configure("TScrollbar", background=c["panel_2"], troughcolor=c["bg"], bordercolor=c["line"], arrowcolor=c["accent_2"])
        style.configure("TLabelframe", background=c["panel"], foreground=c["text"], bordercolor=c["line"])
        style.configure("TLabelframe.Label", background=c["panel"], foreground=c["accent_2"], font=self.font_bold)
        style.configure("Value.TLabel", background=c["panel"], foreground=c["accent_2"], font=self.font_bold)
        style.configure("Muted.TLabel", background=c["panel"], foreground=c["muted"], font=self.font_small)
        style.configure("Head.TLabel", background=c["panel_2"], foreground=c["accent_2"], font=self.font_small)
        self.root.option_add("*TCombobox*Listbox.background", c["card_alt"])
        self.root.option_add("*TCombobox*Listbox.foreground", c["text"])
        self.root.option_add("*TCombobox*Listbox.selectBackground", c["accent"])
        self.root.option_add("*TCombobox*Listbox.selectForeground", c["bg"])

    def build_dashboard(self):
        c = self.colors
        self.root.configure(bg=c["bg"])
        top = tk.Frame(self.root, bg=c["panel"], padx=18, pady=14)
        top.pack(fill=tk.X)
        title_box = tk.Frame(top, bg=c["panel"])
        title_box.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Label(title_box, text="기계 시트 표준 견적 입력 시스템", bg=c["panel"], fg=c["text"], font=(self.font_family, 20, "bold")).pack(anchor="w")
        self.summary_var = tk.StringVar()
        tk.Label(title_box, textvariable=self.summary_var, bg=c["panel"], fg=c["muted"]).pack(anchor="w", pady=(4, 0))
        tk.Label(title_box, text="[카드] 클릭하면 팝업창에 항목 정보 전체와 상세 입력이 표시됩니다.", bg=c["panel"], fg=c["accent_2"], font=self.font_small).pack(anchor="w", pady=(4, 0))
        actions = tk.Frame(top, bg=c["panel"])
        actions.pack(side=tk.RIGHT)
        ttk.Button(actions, text="기계 시트 업로드", command=self.upload_excel_file).pack(side=tk.LEFT, padx=4)
        ttk.Button(actions, text="새 카드 추가", command=self.add_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(actions, text="날짜별 누적 저장", command=self.save_to_excel).pack(side=tk.LEFT, padx=4)
        search = tk.Frame(self.root, bg=c["bg"], padx=16, pady=12)
        search.pack(fill=tk.X)
        tk.Label(search, text="Search", bg=c["bg"], fg=c["text"], font=self.font_bold).pack(side=tk.LEFT, padx=(0, 8))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *args: self.schedule_refresh())
        ttk.Entry(search, textvariable=self.search_var, width=42).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(search, text="검색 초기화", command=self.clear_search).pack(side=tk.LEFT)
        tk.Label(search, text="Search는 품번/품명/기종 대상, 공백/쉼표로 여러 조건 입력 가능", bg=c["bg"], fg=c["muted"]).pack(side=tk.LEFT, padx=(12, 0))
        extra_actions = tk.Frame(search, bg=c["bg"])
        extra_actions.pack(side=tk.RIGHT)
        tk.Label(extra_actions, text="보조 기능", bg=c["bg"], fg=c["muted"], font=self.font_small).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(extra_actions, text="신규 입력 다운로드", command=self.add_download_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(extra_actions, text="전체 선택", command=self.select_visible_items).pack(side=tk.LEFT, padx=4)
        ttk.Button(extra_actions, text="선택 해제", command=self.clear_selection).pack(side=tk.LEFT, padx=4)
        ttk.Button(extra_actions, text="선택 다운로드", command=self.export_selected_items).pack(side=tk.LEFT, padx=4)
        self.card_canvas = tk.Canvas(self.root, bg=c["bg"], highlightthickness=0)
        self.card_scrollbar = ttk.Scrollbar(self.root, orient=tk.VERTICAL, command=self.card_canvas.yview)
        self.card_canvas.configure(yscrollcommand=self.card_scrollbar.set)
        self.card_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.card_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.card_container = tk.Frame(self.card_canvas, bg=c["bg"], padx=16, pady=6)
        self.card_window = self.card_canvas.create_window((0, 0), window=self.card_container, anchor="nw")
        self.card_container.bind("<Configure>", lambda e: self._update_canvas())
        self.card_canvas.bind("<Configure>", lambda e: self._update_canvas())
        self.card_canvas.bind("<Enter>", lambda e: self.card_canvas.bind_all("<MouseWheel>", self._on_mousewheel))
        self.card_canvas.bind("<Leave>", lambda e: self.card_canvas.unbind_all("<MouseWheel>"))

    def _update_canvas(self):
        self.card_canvas.configure(scrollregion=self.card_canvas.bbox("all"))
        self.card_canvas.itemconfigure(self.card_window, width=self.card_canvas.winfo_width())

    def _on_mousewheel(self, event):
        self.card_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def create_blank_item(self, no):
        return {"no": no, "excel_row": None, "created_at": datetime.now().strftime("%Y-%m-%d"), "added_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "source_file": "", "source_month": "", "save_pending": True, "part_no": "", "part_name": "", "model": "", "comment": "", "possible": "가능", "qty": 1, "material": "", "size": "", "m_5axis": 0.0, "m_4axis": 0.0, "m_3axis": 0.0, "m_lathe": 0.0, "m_general": 0.0, "m_finish": 0.0, "m_cmm": 0.0, "m_grind": 0.0, "m_jig": 0.0, "m_prog": 0.0}

    def get_next_no(self):
        return max([item["no"] for item in self.data], default=0) + 1

    def safe_float(self, value):
        if value in (None, ""):
            return 0.0
        if isinstance(value, str) and value.strip().startswith("="):
            return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def safe_int(self, value, default=1):
        if value in (None, ""):
            return default
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default

    def get_session_state_path(self):
        base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
        return os.path.join(base, "MachineEstimate", "session_state.json")

    def save_session_state(self):
        # 엑셀 양식은 입력/출력용으로만 쓰고, 세션 데이터는 이 JSON 파일에 별도로 저장한다.
        state_path = self.get_session_state_path()
        payload = {
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "rates": self.rates,
            "selected_nos": sorted(self.selected_nos),
            "search": self.search_var.get() if hasattr(self, "search_var") else "",
            "items": [item for item in self.data if self.has_item_data(item)],
        }
        try:
            os.makedirs(os.path.dirname(state_path), exist_ok=True)
            with open(state_path, "w", encoding="utf-8") as state_file:
                json.dump(payload, state_file, ensure_ascii=False, indent=2)
        except OSError:
            pass

    def restore_session_state(self):
        # 프로그램 실행 시 이전에 저장된 JSON 세션으로 화면 상태를 완전히 덮어써서 복원한다.
        try:
            with open(self.get_session_state_path(), "r", encoding="utf-8") as state_file:
                payload = json.load(state_file)
        except (OSError, ValueError):
            return
        restored = []
        for raw_item in payload.get("items", []):
            item = self.create_blank_item(raw_item.get("no", 0))
            item.update(raw_item)
            restored.append(item)
        self.data = restored
        self.rates.update(payload.get("rates", {}))
        self.selected_nos = set(payload.get("selected_nos", []))
        if hasattr(self, "search_var"):
            self.search_var.set(payload.get("search", ""))
        self.session_restored_count = len(restored)
        self.session_saved_at = payload.get("saved_at")

    def schedule_refresh(self):
        if self.search_after_id:
            self.root.after_cancel(self.search_after_id)
        self.search_after_id = self.root.after(250, lambda: self.refresh_table(True))

    def show_more_cards(self):
        self.display_limit += self.display_page_size
        self.refresh_table(False)

    def get_filtered_items(self):
        query = self.search_var.get().strip() if hasattr(self, "search_var") else ""
        all_items = [item for item in self.data if self.has_item_data(item)]
        # 최근에 업로드/작성된 카드가 맨 위로 오도록 added_at 내림차순 정렬한다.
        all_items.sort(key=lambda item: item.get("added_at", ""), reverse=True)
        visible_items = [item for item in all_items if self.item_matches_search(item, query)]
        return all_items, visible_items

    def read_cards_from_workbook(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        if "기계" not in wb.sheetnames:
            raise ValueError("선택한 파일에서 '기계' 시트를 찾을 수 없습니다.")
        ws = wb["기계"]
        mach_keys = ["m_5axis", "m_4axis", "m_3axis", "m_lathe", "m_general", "m_finish", "m_cmm", "m_grind", "m_jig", "m_prog"]
        cards = []
        summary_row = self.find_summary_row(ws)
        abs_path = os.path.abspath(file_path)
        created_at = datetime.fromtimestamp(os.path.getmtime(abs_path)).strftime("%Y-%m-%d")
        source_month = os.path.basename(os.path.dirname(abs_path))
        for row in range(7, summary_row):
            if not self.row_has_input_data(ws, row):
                continue
            item = self.create_blank_item(len(cards) + 1)
            item["created_at"] = created_at
            item["source_file"] = os.path.basename(abs_path)
            item["source_month"] = source_month
            item["save_pending"] = True
            item["part_no"] = str(ws.cell(row=row, column=2).value or "")
            item["part_name"] = str(ws.cell(row=row, column=3).value or "")
            item["comment"] = str(ws.cell(row=row, column=4).value or "")
            item["possible"] = str(ws.cell(row=row, column=5).value or "가능")
            item["qty"] = self.safe_int(ws.cell(row=row, column=6).value, 1)
            item["material"] = str(ws.cell(row=row, column=7).value or "")
            size_parts = [ws.cell(row=row, column=col).value for col in range(8, 11)]
            item["size"] = " x ".join(str(v) for v in size_parts if v not in (None, ""))
            for col, key in zip(range(11, 21), mach_keys):
                item[key] = self.safe_float(ws.cell(row=row, column=col).value)
            cards.append(item)
        return cards

    def upload_excel_file(self):
        file_path = filedialog.askopenfilename(title="업로드할 견적 양식 선택", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_path:
            return
        try:
            cards = self.read_cards_from_workbook(file_path)
        except Exception as exc:
            messagebox.showerror("업로드 오류", f"선택한 파일을 불러오지 못했습니다.\n\n{exc}")
            return
        if not cards:
            messagebox.showwarning("업로드 안내", "선택한 파일에서 입력된 견적 항목을 찾지 못했습니다.")
            return
        if self.data and not messagebox.askyesno("업로드 확인", f"선택한 파일에서 {len(cards)}개 항목을 찾았습니다.\n현재 카드 목록에 추가하시겠습니까?"):
            return
        next_no = self.get_next_no()
        for offset, item in enumerate(cards):
            item["no"] = next_no + offset
            item["excel_row"] = None
            item["save_pending"] = True
            self.data.append(item)
        self.save_session_state()
        self.refresh_table(True)
        messagebox.showinfo("업로드 완료", f"{len(cards)}개 항목을 카드로 불러왔습니다.\n저장하면 오늘 날짜 누적 파일에 추가됩니다.")

    def clear_search(self):
        self.search_var.set("")
        self.refresh_table(True)

    def item_matches_search(self, item, query):
        if not query:
            return True
        terms = [term.lower() for term in re.split(r"[\s,]+", query.strip()) if term]
        # 검색 범위는 품번/품명/기종 3종류로 한정한다(요청에 따라 Material/Size/Coment/금액 등은 제외).
        values = [item.get("part_no", ""), item.get("part_name", ""), item.get("model", "")]
        haystack = " ".join(str(v).lower() for v in values)
        return all(term in haystack for term in terms)

    def toggle_item_selection(self, no, is_selected):
        if is_selected:
            self.selected_nos.add(no)
        else:
            self.selected_nos.discard(no)
        self.refresh_table(False)

    def select_visible_items(self):
        _, visible_items = self.get_filtered_items()
        for item in visible_items:
            self.selected_nos.add(item["no"])
        self.refresh_table(False)

    def clear_selection(self):
        self.selected_nos.clear()
        self.refresh_table(False)

    def get_status_colors(self, status):
        if status == "불가":
            return self.colors["danger_bg"], self.colors["danger_fg"], "#e05561"
        if status == "검토필요":
            return self.colors["warn_bg"], self.colors["warn_fg"], "#d88a4f"
        return self.colors["success_bg"], self.colors["success_fg"], "#3ecf8e"

    def calc_row(self, item):
        times = [item["m_5axis"], item["m_4axis"], item["m_3axis"], item["m_lathe"], item["m_general"], item["m_finish"], item["m_cmm"], item["m_grind"], item["m_jig"], item["m_prog"]]
        sum_time = sum(times)
        cost_sum = sum(item[key] * self.rates[key] for key in self.rates)
        unit_price = cost_sum * item["qty"]
        final_price = int(unit_price // 1000 * 1000)
        return sum_time, unit_price, final_price

    def refresh_table(self, reset_limit=False):
        if reset_limit:
            self.display_limit = self.display_page_size
        self.search_after_id = None
        for widget in self.card_container.winfo_children():
            widget.destroy()
        query = self.search_var.get().strip() if hasattr(self, "search_var") else ""
        all_items, visible_items = self.get_filtered_items()
        self.selected_nos.intersection_update({item["no"] for item in all_items})
        selected_count = len([item for item in all_items if item["no"] in self.selected_nos])
        total_amount = sum(self.calc_row(item)[2] for item in all_items)
        today = datetime.now().strftime("%Y-%m-%d")
        search_note = f"  |  검색 결과 {len(visible_items)}건" if query else ""
        select_note = f"  |  선택 {selected_count}건"
        if self.session_saved_at:
            session_note = f"  |  세션 복원 {self.session_restored_count}건 (저장 {self.session_saved_at})"
        else:
            session_note = "  |  새 세션 (저장된 데이터 없음)"
        self.summary_var.set(f"작성일 {today}  |  카드 {len(all_items)}건  |  총 견적금액 {total_amount:,}원{search_note}{select_note}{session_note}")
        if not all_items:
            self.render_empty_message("아직 작성된 견적 카드가 없습니다.", "상단의 '기계 시트 업로드' 또는 '새 카드 추가'를 사용하세요.")
            return
        if not visible_items:
            self.render_empty_message("검색 결과가 없습니다.", "품번, 품명, 기종으로 검색할 수 있습니다.")
            return
        for item in visible_items[:self.display_limit]:
            self.render_card(item)
        if self.display_limit < len(visible_items):
            more = tk.Frame(self.card_container, bg=self.colors["bg"], padx=8, pady=14)
            more.pack(fill=tk.X)
            ttk.Button(more, text=f"더 보기 ({self.display_limit}/{len(visible_items)})", command=self.show_more_cards).pack(anchor="center")

    def render_empty_message(self, title, detail):
        c = self.colors
        empty = tk.Frame(self.card_container, bg=c["card"], highlightthickness=1, highlightbackground=c["line"], padx=28, pady=28)
        empty.pack(fill=tk.X, pady=10)
        tk.Label(empty, text=title, bg=c["card"], fg=c["text"], font=(self.font_family, 17, "bold")).pack(anchor="w")
        tk.Label(empty, text=detail, bg=c["card"], fg=c["muted"]).pack(anchor="w", pady=(8, 0))

    def render_selection_checkbox(self, parent, no, checked):
        # 기본 tk 체크박스 표시기(Windows에서 ~13px, 폰트를 키워도 커지지 않음)의 2배인
        # 26px 정사각형을 Frame(pack_propagate 고정)으로 직접 그려서 클릭 토글한다.
        c = self.colors
        size = 26
        box_bg = c["accent"] if checked else c["card_alt"]
        box_fg = c["bg"] if checked else c["muted"]
        holder = tk.Frame(parent, width=size, height=size, bg=box_bg, highlightthickness=1, highlightbackground=c["line"], cursor="hand2")
        holder.pack_propagate(False)
        glyph = tk.Label(holder, text=("V" if checked else ""), bg=box_bg, fg=box_fg, font=(self.font_family, 13, "bold"))
        glyph.pack(expand=True)
        toggle = lambda event, no=no, checked=checked: self.toggle_item_selection(no, not checked)
        holder.bind("<Button-1>", toggle)
        glyph.bind("<Button-1>", toggle)
        holder.is_control = True
        return holder

    def render_card(self, item):
        c = self.colors
        sum_time, _, final_price = self.calc_row(item)
        status_bg, status_fg, border = self.get_status_colors(item["possible"])
        card = tk.Frame(self.card_container, bg=c["card"], highlightthickness=1, highlightbackground=border, padx=16, pady=14)
        card.pack(fill=tk.X, pady=8)
        card.columnconfigure(1, weight=1)
        tk.Frame(card, bg=status_fg, width=7).grid(row=0, column=0, rowspan=4, sticky="nsw", padx=(0, 14))
        head = tk.Frame(card, bg=c["card"])
        head.grid(row=0, column=1, sticky="ew")
        head.columnconfigure(2, weight=1)
        self.render_selection_checkbox(head, item["no"], item["no"] in self.selected_nos).grid(row=0, column=0, rowspan=2, sticky="w", padx=(0, 8))
        tk.Label(head, text="선택", bg=c["card"], fg=c["muted"], font=self.font_small).grid(row=0, column=1, rowspan=2, sticky="w", padx=(0, 12))
        title_box = tk.Frame(head, bg=c["card"])
        title_box.grid(row=0, column=2, rowspan=2, sticky="w")
        title_row = tk.Frame(title_box, bg=c["card"])
        title_row.pack(anchor="w")
        tk.Label(title_row, text=item["part_no"] or f"NO. {item['no']} 미입력 카드", bg=c["card"], fg=c["text"], font=(self.font_family, 17, "bold")).pack(side=tk.LEFT)
        if item.get("save_pending"):
            tk.Label(title_row, text="NEW", bg=c["new_bg"], fg=c["new_fg"], font=self.font_small, padx=6, pady=1, highlightthickness=1, highlightbackground=c["new_border"]).pack(side=tk.LEFT, padx=(8, 0))
        tk.Label(title_box, text=item["part_name"] or "품명 미입력", bg=c["card"], fg=c["muted"], font=self.font_bold).pack(anchor="w", pady=(2, 0))
        tk.Label(head, text=item["possible"], bg=status_bg, fg=status_fg, font=self.font_bold, padx=12, pady=4).grid(row=0, column=3, rowspan=2, sticky="e", padx=(10, 0))
        body = tk.Frame(card, bg=c["card"])
        body.grid(row=1, column=1, sticky="ew", pady=(12, 0))
        # 기계 시트 대상 엑셀 파일 정보는 메인 화면에 노출하지 않고 팝업창에서만 표기한다.
        facts = [("기종", item["model"] or "-"), ("작성일", item["created_at"]), ("Qty", item["qty"]), ("Material", item["material"] or "-"), ("Size", item["size"] or "-"), ("시간합계", f"{sum_time:.1f} h"), ("최종단가", f"{final_price:,} 원")]
        for col in range(len(facts)):
            body.columnconfigure(col, weight=1)
        for idx, (label, value) in enumerate(facts):
            box = tk.Frame(body, bg=c["card_alt"], padx=10, pady=8)
            box.grid(row=0, column=idx, sticky="nsew", padx=3)
            tk.Label(box, text=label, bg=c["card_alt"], fg=c["muted"], font=self.font_small).pack(anchor="w")
            tk.Label(box, text=str(value), bg=c["card_alt"], fg=c["text"], wraplength=220, justify="left").pack(anchor="w", pady=(2, 0))
        foot = tk.Frame(card, bg=c["card"])
        foot.grid(row=2, column=1, sticky="ew", pady=(12, 0))
        tk.Label(foot, text=item["comment"] or "메모 없음", bg=c["card"], fg=c["muted"], wraplength=920, justify="left").pack(side=tk.LEFT)
        ttk.Button(foot, text="수정", command=lambda no=item["no"]: self.open_popup(no)).pack(side=tk.RIGHT)
        self.bind_card_click(card, item["no"])

    def bind_card_click(self, widget, no):
        # 카드 안쪽 어디를 클릭해도 팝업이 열리도록 하위 위젯까지 재귀로 연결한다.
        # is_control로 표시된 위젯(예: 커스텀 체크박스)은 자체 클릭 동작을 덮어쓰지 않도록 건너뛴다.
        if getattr(widget, "is_control", False):
            return
        if not isinstance(widget, (ttk.Button, tk.Button, tk.Checkbutton, ttk.Checkbutton)):
            widget.bind("<Button-1>", lambda event, card_no=no: self.open_popup(card_no))
        for child in widget.winfo_children():
            self.bind_card_click(child, no)
    def get_app_dir(self):
        if getattr(sys, "frozen", False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.abspath(__file__))

    def parse_version(self, text):
        match = re.search(r"(\d+)\.(\d+)\.(\d+)", str(text))
        if not match:
            return None
        return tuple(int(part) for part in match.groups())

    def get_update_state_path(self):
        base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
        return os.path.join(base, "MachineEstimate", "update_state.txt")

    def read_applied_updates(self):
        try:
            with open(self.get_update_state_path(), "r", encoding="utf-8") as state_file:
                return {line.strip() for line in state_file if line.strip()}
        except OSError:
            return set()

    def mark_update_applied(self, marker):
        state_path = self.get_update_state_path()
        try:
            os.makedirs(os.path.dirname(state_path), exist_ok=True)
            with open(state_path, "a", encoding="utf-8") as state_file:
                state_file.write(marker + "\n")
        except OSError:
            pass

    def check_for_update(self):
        update_dir = os.path.join(self.get_app_dir(), "update")
        if not os.path.isdir(update_dir):
            return
        current_version = self.parse_version(APP_VERSION) or (0, 0, 0)
        applied = self.read_applied_updates()
        candidates = []
        for name in os.listdir(update_dir):
            if not name.lower().endswith(".exe"):
                continue
            path = os.path.join(update_dir, name)
            try:
                marker = f"{name}|{int(os.path.getmtime(path))}"
            except OSError:
                continue
            # 같은 파일로 업데이트 안내가 반복되지 않도록 버전과 적용 이력을 함께 확인한다.
            file_version = self.parse_version(name)
            if file_version is not None and file_version <= current_version:
                continue
            if marker in applied:
                continue
            candidates.append((os.path.getmtime(path), name, path, marker))
        if not candidates:
            return
        candidates.sort(reverse=True)
        _, latest_name, installer_path, marker = candidates[0]
        if not messagebox.askyesno("업데이트 확인", f"update 폴더에서 새 설치 파일을 발견했습니다.\n\n현재 버전 v{APP_VERSION}\n설치 파일 {latest_name}\n\n지금 실행해서 업데이트하시겠습니까?\n(설치 중 현재 실행 중인 프로그램은 자동으로 종료 후 재시작됩니다.)"):
            self.mark_update_applied(marker)
            return
        try:
            subprocess.Popen([installer_path])
            self.mark_update_applied(marker)
        except OSError as exc:
            messagebox.showerror("업데이트 실행 오류", f"설치 파일을 실행할 수 없습니다.\n{exc}")

    def get_resource_path(self, filename):
        candidates = [os.path.join(self.get_app_dir(), filename), os.path.join(os.getcwd(), filename), os.path.join(os.getcwd(), "견적_산정", "양식", filename)]
        if getattr(sys, "frozen", False):
            candidates.insert(1, os.path.join(getattr(sys, "_MEIPASS", ""), filename))
        for path in candidates:
            if path and os.path.exists(path):
                return path
        return candidates[0]

    def get_estimate_root_dir(self):
        app_dir = self.get_app_dir()
        candidates = [os.path.join(app_dir, "견적_산정"), os.path.join(os.path.dirname(app_dir), "견적_산정"), os.path.join(os.getcwd(), "견적_산정")]
        for path in candidates:
            if os.path.isdir(path):
                return path
        return candidates[0]

    def get_dated_output_path(self, create_dir=True):
        today = datetime.now()
        output_dir = os.path.join(self.get_estimate_root_dir(), f"{today.year}년도", f"{today.month:02d}월")
        if create_dir:
            os.makedirs(output_dir, exist_ok=True)
        return os.path.join(output_dir, f"견적누적_{today:%Y-%m-%d}.xlsx")

    def row_has_input_data(self, ws, row):
        for col in [2, 3, 4, 5, 6, 7, 8] + list(range(11, 21)):
            if ws.cell(row=row, column=col).value not in (None, ""):
                return True
        return False

    def find_summary_row(self, ws):
        for row in range(7, ws.max_row + 1):
            if ws.cell(row=row, column=22).value == "합계":
                return row
        return max(ws.max_row + 1, 14)

    def find_next_input_row(self, ws):
        summary_row = self.find_summary_row(ws)
        for row in range(7, summary_row):
            if not self.row_has_input_data(ws, row):
                return row
        ws.insert_rows(summary_row)
        self.copy_row_style(ws, max(7, summary_row - 1), summary_row)
        return summary_row

    def copy_row_style(self, ws, source_row, target_row):
        for col in range(1, 24):
            source = ws.cell(row=source_row, column=col)
            target = ws.cell(row=target_row, column=col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.border:
                target.border = copy(source.border)
            if source.fill:
                target.fill = copy(source.fill)
            if source.font:
                target.font = copy(source.font)
            if source.protection:
                target.protection = copy(source.protection)
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    def apply_row_formulas(self, ws, row):
        ws.cell(row=row, column=1, value=f"=ROW()-6")
        ws.cell(row=row, column=21, value=f"=SUM($K{row}:$T{row})")
        ws.cell(row=row, column=22, value=f"=(K{row}*$K$5+L{row}*$L$5+M{row}*$M$5+N{row}*$N$5+O{row}*$O$5+P{row}*$P$5+Q{row}*$Q$5+T{row}*$T$5+R{row}*$R$5+$S$5*S{row})*$F{row}")
        ws.cell(row=row, column=23, value=f"=ROUNDDOWN($V{row},-3)")

    def update_summary_formula(self, ws):
        summary_row = self.find_summary_row(ws)
        if summary_row > 7:
            ws.cell(row=summary_row, column=22, value="합계")
            ws.cell(row=summary_row, column=23, value=f"=SUM(W7:W{summary_row - 1})")

    def has_item_data(self, item):
        has_text = any(str(item.get(key, "")).strip() for key in ["part_no", "part_name", "comment", "material", "size"])
        has_time = any(float(item.get(key, 0) or 0) > 0 for key in ["m_5axis", "m_4axis", "m_3axis", "m_lathe", "m_general", "m_finish", "m_cmm", "m_grind", "m_jig", "m_prog"])
        return has_text or has_time

    def add_row(self):
        new_no = self.get_next_no()
        self.data.append(self.create_blank_item(new_no))
        self.refresh_table(True)
        self.open_popup(new_no)

    def add_download_row(self):
        new_no = self.get_next_no()
        self.data.append(self.create_blank_item(new_no))
        self.refresh_table(True)
        self.open_popup(new_no, export_on_save=True)
    def build_popup_info(self, parent, item):
        # 카드 클릭 시 항목 정보 전체를 팝업창 상단에 표기한다(기계 시트 대상 엑셀 파일 포함).
        info_box = ttk.LabelFrame(parent, text="항목 정보", padding=10)
        info_box.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        sum_time, unit_price, final_price = self.calc_row(item)
        summary_vars = {
            "sum_time": tk.StringVar(value=f"{sum_time:.1f} h"),
            "unit_price": tk.StringVar(value=f"{int(unit_price):,} 원"),
            "final_price": tk.StringVar(value=f"{final_price:,} 원"),
        }
        static_facts = [
            ("NO", str(item["no"])),
            ("작성일", item["created_at"]),
            ("저장 상태", "저장 대기" if item.get("save_pending") else "저장 완료"),
            ("기계 시트 폴더", item["source_month"] or "신규"),
            ("기계 시트 파일", item["source_file"] or "저장 대기"),
            ("엑셀 행", str(item["excel_row"]) if item.get("excel_row") else "-"),
        ]
        live_facts = [("시간합계", summary_vars["sum_time"]), ("단가합계", summary_vars["unit_price"]), ("최종 견적금액", summary_vars["final_price"])]
        total_cols = len(static_facts) + len(live_facts)
        for col in range(total_cols):
            info_box.columnconfigure(col, weight=1)
        for idx, (label, value) in enumerate(static_facts):
            tile = ttk.Frame(info_box, padding=6)
            tile.grid(row=0, column=idx, sticky="nsew", padx=3)
            ttk.Label(tile, text=label, style="Muted.TLabel").pack(anchor="w")
            ttk.Label(tile, text=value, wraplength=170, justify="left").pack(anchor="w", pady=(2, 0))
        for offset, (label, var) in enumerate(live_facts):
            tile = ttk.Frame(info_box, padding=6)
            tile.grid(row=0, column=len(static_facts) + offset, sticky="nsew", padx=3)
            ttk.Label(tile, text=label, style="Muted.TLabel").pack(anchor="w")
            ttk.Label(tile, textvariable=var, style="Value.TLabel").pack(anchor="w", pady=(2, 0))
        return summary_vars

    def open_popup(self, no, export_on_save=False):
        item = next((row for row in self.data if row["no"] == no), None)
        if not item:
            return
        opened = getattr(self, "open_popups", {})
        existing = opened.get(no)
        if existing is not None and existing.winfo_exists():
            existing.lift()
            existing.focus_force()
            return
        popup = tk.Toplevel(self.root)
        self.open_popups = opened
        self.open_popups[no] = popup
        popup.title(f"[NO. {no}] 기계 시트 항목 입력")
        # 화면 배율에 따라 팝업이 화면 밖으로 넘치지 않도록 표시 영역에 맞춰 크기를 정한다.
        width = min(1180, max(900, popup.winfo_screenwidth() - 80))
        height = min(800, max(600, popup.winfo_screenheight() - 120))
        pos_x = max(0, (popup.winfo_screenwidth() - width) // 2)
        pos_y = max(0, (popup.winfo_screenheight() - height) // 3)
        popup.geometry(f"{width}x{height}+{pos_x}+{pos_y}")
        popup.minsize(min(900, width), min(600, height))
        popup.configure(bg=self.colors["bg"])
        popup.grab_set()
        popup.bind("<Destroy>", lambda event, card_no=no: self.open_popups.pop(card_no, None) if event.widget is popup else None)

        frame = ttk.Frame(popup, padding=16)
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(0, weight=3)
        frame.columnconfigure(1, weight=2)

        summary_vars = self.build_popup_info(frame, item)

        left = ttk.Frame(frame)
        left.grid(row=1, column=0, sticky="nsew", padx=(0, 18))
        fields = {}

        info = ttk.LabelFrame(left, text="기본 정보", padding=10)
        info.pack(fill=tk.X, pady=(0, 10))
        info.columnconfigure(1, weight=1)
        info.columnconfigure(3, weight=1)

        ttk.Label(info, text="품번 *").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        fields["part_no"] = tk.StringVar(value=item["part_no"])
        ttk.Entry(info, textvariable=fields["part_no"], width=18).grid(row=0, column=1, sticky="ew", padx=6, pady=6)

        ttk.Label(info, text="품명 *").grid(row=0, column=2, sticky="e", padx=6, pady=6)
        fields["part_name"] = tk.StringVar(value=item["part_name"])
        ttk.Entry(info, textvariable=fields["part_name"], width=18).grid(row=0, column=3, sticky="ew", padx=6, pady=6)

        ttk.Label(info, text="기종").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        fields["model"] = tk.StringVar(value=item["model"])
        ttk.Entry(info, textvariable=fields["model"], width=18).grid(row=1, column=1, sticky="ew", padx=6, pady=6)

        ttk.Label(info, text="Material").grid(row=1, column=2, sticky="e", padx=6, pady=6)
        fields["material"] = tk.StringVar(value=item["material"])
        ttk.Entry(info, textvariable=fields["material"], width=18).grid(row=1, column=3, sticky="ew", padx=6, pady=6)

        ttk.Label(info, text="수량(Qty)").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        fields["qty"] = tk.StringVar(value=str(item["qty"]))
        ttk.Entry(info, textvariable=fields["qty"], width=18).grid(row=2, column=1, sticky="ew", padx=6, pady=6)

        ttk.Label(info, text="가능여부").grid(row=2, column=2, sticky="e", padx=6, pady=6)
        fields["possible"] = tk.StringVar(value=item["possible"])
        ttk.Combobox(info, textvariable=fields["possible"], values=["가능", "불가", "검토필요"], state="readonly", width=15).grid(row=2, column=3, sticky="w", padx=6, pady=6)

        ttk.Label(info, text="Comment").grid(row=3, column=0, sticky="e", padx=6, pady=6)
        fields["comment"] = tk.StringVar(value=item["comment"])
        ttk.Entry(info, textvariable=fields["comment"], width=18).grid(row=3, column=1, columnspan=3, sticky="ew", padx=6, pady=6)

        size_box = ttk.LabelFrame(left, text="Size (소재 규격) 입력", padding=10)
        size_box.pack(fill=tk.X)

        shape_var = tk.StringVar(value="블록")
        t_var = tk.StringVar()
        w_var = tk.StringVar()
        l_var = tk.StringVar()
        d_var = tk.StringVar()
        custom_size_var = tk.StringVar()
        final_size_var = tk.StringVar(value=item["size"])

        current_size = item["size"].strip()
        if "Ø" in current_size or "D*" in current_size:
            shape_var.set("로드")
            matches = re.findall(r"[\d\.]+", current_size)
            if len(matches) >= 2:
                d_var.set(matches[0])
                l_var.set(matches[1])
        elif "T" in current_size or "W" in current_size or "*" in current_size:
            shape_var.set("블록")
            matches = re.findall(r"[\d\.]+", current_size)
            if len(matches) >= 3:
                t_var.set(matches[0])
                w_var.set(matches[1])
                l_var.set(matches[2])
        elif current_size:
            shape_var.set("직접입력")
            custom_size_var.set(current_size)

        radio_frame = ttk.Frame(size_box)
        radio_frame.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(radio_frame, text="형상 선택").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Radiobutton(radio_frame, text="블록 (T x W x L)", value="블록", variable=shape_var).pack(side=tk.LEFT, padx=6)
        ttk.Radiobutton(radio_frame, text="로드/원봉 (D x L)", value="로드", variable=shape_var).pack(side=tk.LEFT, padx=6)
        ttk.Radiobutton(radio_frame, text="직접 입력", value="직접입력", variable=shape_var).pack(side=tk.LEFT, padx=6)

        input_subframe = ttk.Frame(size_box, padding=5)
        input_subframe.pack(fill=tk.X)

        def update_size_preview(*args):
            shape = shape_var.get()
            if shape == "블록":
                t_val = t_var.get().strip()
                w_val = w_var.get().strip()
                l_val = l_var.get().strip()
                if t_val or w_val or l_val:
                    if t_val and w_val and l_val:
                        final_size_var.set(f"{t_val}T * {w_val}W * {l_val}L")
                    else:
                        final_size_var.set(f"{t_val} * {w_val} * {l_val}")
                else:
                    final_size_var.set("")
            elif shape == "로드":
                d_val = d_var.get().strip()
                l_val = l_var.get().strip()
                if d_val or l_val:
                    if d_val and l_val:
                        final_size_var.set(f"Ø{d_val} * {l_val}L")
                    else:
                        final_size_var.set(f"Ø{d_val} * {l_val}")
                else:
                    final_size_var.set("")
            else:
                final_size_var.set(custom_size_var.get().strip())

        for variable in (t_var, w_var, l_var, d_var, custom_size_var):
            variable.trace_add("write", update_size_preview)

        def render_size_inputs(*args):
            for widget in input_subframe.winfo_children():
                widget.destroy()
            shape = shape_var.get()
            if shape == "블록":
                ttk.Label(input_subframe, text="T (두께)").grid(row=0, column=0, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=t_var, width=10).grid(row=0, column=1, padx=5, pady=2)
                ttk.Label(input_subframe, text="W (폭)").grid(row=0, column=2, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=w_var, width=10).grid(row=0, column=3, padx=5, pady=2)
                ttk.Label(input_subframe, text="L (길이)").grid(row=0, column=4, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=l_var, width=10).grid(row=0, column=5, padx=5, pady=2)
            elif shape == "로드":
                ttk.Label(input_subframe, text="D (외경/Ø)").grid(row=0, column=0, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=d_var, width=12).grid(row=0, column=1, padx=5, pady=2)
                ttk.Label(input_subframe, text="L (길이)").grid(row=0, column=2, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=l_var, width=12).grid(row=0, column=3, padx=5, pady=2)
            else:
                ttk.Label(input_subframe, text="Size 규격").grid(row=0, column=0, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=custom_size_var, width=35).grid(row=0, column=1, padx=5, pady=2)
            update_size_preview()

        shape_var.trace_add("write", render_size_inputs)
        render_size_inputs()

        preview = ttk.Frame(size_box)
        preview.pack(fill=tk.X, pady=(4, 0))
        ttk.Label(preview, text="최종 반영 Size", style="Muted.TLabel").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(preview, textvariable=final_size_var, style="Value.TLabel").pack(side=tk.LEFT)

        mach_keys = [("m_5axis", "5축"), ("m_4axis", "4축"), ("m_3axis", "3축"), ("m_lathe", "선반"), ("m_general", "범용"), ("m_finish", "사상"), ("m_cmm", "3차원"), ("m_grind", "연마"), ("m_jig", "지그"), ("m_prog", "프로그래밍")]
        times = ttk.LabelFrame(frame, text="각 공정별 시간 / 단가 입력", padding=10)
        times.grid(row=1, column=1, sticky="nsew")
        times.columnconfigure(1, weight=1)
        times.columnconfigure(2, weight=1)
        times.columnconfigure(3, weight=1)
        rate_fields = {}
        amount_vars = {}
        for col, header in enumerate(["공정", "시간(h)", "단가(원)", "금액(원)"]):
            ttk.Label(times, text=header, style="Head.TLabel", padding=(6, 4)).grid(row=0, column=col, sticky="ew", padx=3, pady=(0, 6))
        for idx, (key, label) in enumerate(mach_keys):
            row_index = idx + 1
            ttk.Label(times, text=label).grid(row=row_index, column=0, sticky="w", padx=6, pady=4)
            fields[key] = tk.StringVar(value=(str(item[key]) if item[key] > 0 else ""))
            ttk.Entry(times, textvariable=fields[key], width=10).grid(row=row_index, column=1, sticky="ew", padx=4, pady=4)
            rate_fields[key] = tk.StringVar(value=str(int(self.rates[key])))
            ttk.Entry(times, textvariable=rate_fields[key], width=10).grid(row=row_index, column=2, sticky="ew", padx=4, pady=4)
            amount_vars[key] = tk.StringVar(value="0")
            ttk.Label(times, textvariable=amount_vars[key], style="Value.TLabel", anchor="e").grid(row=row_index, column=3, sticky="ew", padx=4, pady=4)
        ttk.Label(times, text="단가는 견적 양식 5행에 공통 적용됩니다.", style="Muted.TLabel").grid(row=len(mach_keys) + 1, column=0, columnspan=4, sticky="w", padx=6, pady=(8, 0))

        def parse_number(text):
            text = text.strip().replace(",", "")
            if not text:
                return 0.0
            try:
                return float(text)
            except ValueError:
                return None

        def update_amounts(*args):
            sum_time, cost_sum, has_error = 0.0, 0.0, False
            for key, _ in mach_keys:
                hours = parse_number(fields[key].get())
                rate = parse_number(rate_fields[key].get())
                if hours is None or rate is None or hours < 0 or rate < 0:
                    amount_vars[key].set("입력오류")
                    has_error = True
                    continue
                amount_vars[key].set(f"{int(hours * rate):,}")
                sum_time += hours
                cost_sum += hours * rate
            qty = parse_number(fields["qty"].get())
            if qty is None or qty <= 0:
                qty = 1
            unit_price = cost_sum * qty
            summary_vars["sum_time"].set("입력오류" if has_error else f"{sum_time:.1f} h")
            summary_vars["unit_price"].set("입력오류" if has_error else f"{int(unit_price):,} 원")
            summary_vars["final_price"].set("입력오류" if has_error else f"{int(unit_price // 1000 * 1000):,} 원")

        for key, _ in mach_keys:
            fields[key].trace_add("write", update_amounts)
            rate_fields[key].trace_add("write", update_amounts)
        fields["qty"].trace_add("write", update_amounts)
        update_amounts()

        def save_and_close(next_item=False):
            if not fields["part_no"].get().strip() or not fields["part_name"].get().strip():
                messagebox.showerror("입력 오류", "품번과 품명은 필수 입력 항목입니다.", parent=popup)
                return
            try:
                qty_value = int(fields["qty"].get().strip())
                if qty_value <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("입력 오류", "수량(Qty)은 1 이상의 정수로 입력하셔야 합니다.", parent=popup)
                return
            mach_values = {}
            rate_values = {}
            for key, label in mach_keys:
                value = fields[key].get().strip()
                try:
                    mach_values[key] = float(value) if value else 0.0
                except ValueError:
                    messagebox.showerror("입력 오류", "공수 시간은 숫자로 입력해 주세요 (예: 1.5).", parent=popup)
                    return
                rate_text = rate_fields[key].get().strip().replace(",", "")
                try:
                    rate_value = float(rate_text) if rate_text else 0.0
                except ValueError:
                    messagebox.showerror("입력 오류", f"'{label}' 단가는 숫자로 입력해 주세요.", parent=popup)
                    return
                if rate_value < 0:
                    messagebox.showerror("입력 오류", f"'{label}' 단가는 0 이상으로 입력해 주세요.", parent=popup)
                    return
                rate_values[key] = rate_value
            item["part_no"] = fields["part_no"].get().strip()
            item["part_name"] = fields["part_name"].get().strip()
            item["model"] = fields["model"].get().strip()
            item["material"] = fields["material"].get().strip()
            item["size"] = final_size_var.get().strip()
            item["comment"] = fields["comment"].get().strip()
            item["possible"] = fields["possible"].get().strip()
            item["qty"] = qty_value
            item["save_pending"] = True
            for key, value in mach_values.items():
                item[key] = value
            self.rates.update(rate_values)
            if export_on_save:
                if not self.export_items_to_excel([item], default_name=f"신규견적_{datetime.now():%Y-%m-%d}.xlsx", parent=popup):
                    return
            self.save_session_state()
            self.refresh_table(True)
            popup.destroy()
            if export_on_save and not messagebox.askyesno("신규 입력 반영", "방금 다운로드한 신규 항목을 현재 카드 목록에도 유지하시겠습니까?", parent=self.root):
                self.data = [row for row in self.data if row["no"] != no]
                self.selected_nos.discard(no)
                self.save_session_state()
                self.refresh_table(True)
                return
            if next_item:
                next_no = no + 1
                if not any(row["no"] == next_no for row in self.data):
                    self.data.append(self.create_blank_item(next_no))
                self.refresh_table(True)
                self.open_popup(next_no)

        btns = ttk.Frame(frame)
        btns.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(18, 0))
        ttk.Button(btns, text="저장 후 다음 항목 입력", command=lambda: save_and_close(True)).pack(side=tk.LEFT)
        ttk.Button(btns, text="저장", command=lambda: save_and_close(False)).pack(side=tk.RIGHT)

    def save_to_excel(self):
        if self.is_saving:
            messagebox.showwarning("저장 진행 중", "이미 저장 중입니다. 잠시만 기다려 주세요.")
            return
        self.is_saving = True
        try:
            self._save_to_excel_impl()
        finally:
            self.is_saving = False

    def _save_items_to_workbook(self, ws, items, update_item_rows):
        mach_cols = [11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
        mach_keys = ["m_5axis", "m_4axis", "m_3axis", "m_lathe", "m_general", "m_finish", "m_cmm", "m_grind", "m_jig", "m_prog"]
        # 입력창에서 수정한 공정별 단가를 양식의 단가 행(5행)에 반영한다.
        for col_idx, key in zip(mach_cols, mach_keys):
            ws.cell(row=5, column=col_idx, value=self.rates[key])
        saved_count = 0
        for item in items:
            row = item.get("excel_row") if update_item_rows else None
            if not row or row >= self.find_summary_row(ws):
                row = self.find_next_input_row(ws)
                if update_item_rows:
                    item["excel_row"] = row
            self.copy_row_style(ws, 7, row)
            self.apply_row_formulas(ws, row)
            ws.cell(row=row, column=2, value=item["part_no"].strip() if item["part_no"] else None)
            ws.cell(row=row, column=3, value=item["part_name"].strip() if item["part_name"] else None)
            ws.cell(row=row, column=4, value=item["comment"].strip() if item["comment"] else None)
            ws.cell(row=row, column=5, value=item["possible"])
            ws.cell(row=row, column=6, value=item["qty"])
            ws.cell(row=row, column=7, value=item["material"].strip() if item["material"] else None)
            ws.cell(row=row, column=8, value=item["size"].strip() if item["size"] else None)
            for col_idx, key in zip(mach_cols, mach_keys):
                val = item[key]
                ws.cell(row=row, column=col_idx, value=val if val > 0 else None)
            if update_item_rows:
                item["save_pending"] = False
            saved_count += 1
        self.update_summary_formula(ws)
        return saved_count

    def _save_to_excel_impl(self):
        template_path = self.get_resource_path("견적용.xlsx")
        if not os.path.exists(template_path):
            messagebox.showerror("파일 오류", f"'{template_path}' 파일이 존재하지 않습니다.")
            return
        save_items = [item for item in self.data if self.has_item_data(item) and (item.get("save_pending") or item.get("excel_row"))]
        if not save_items:
            messagebox.showwarning("저장 대상 없음", "입력된 항목이 없어 저장하지 않았습니다.")
            return
        output_path = self.get_dated_output_path(True)
        source_path = output_path if os.path.exists(output_path) else template_path
        wb = openpyxl.load_workbook(source_path)
        if "기계" not in wb.sheetnames:
            messagebox.showerror("시트 오류", "'기계' 시트를 찾을 수 없습니다.")
            return
        ws = wb["기계"]
        saved_count = self._save_items_to_workbook(ws, save_items, True)
        wb.save(output_path)
        for item in save_items:
            item["source_file"] = os.path.basename(output_path)
            item["source_month"] = os.path.basename(os.path.dirname(output_path))
        self.save_session_state()
        messagebox.showinfo("누적 저장 완료", f"{saved_count}개 항목을 날짜별 파일에 누적 저장했습니다.\n\n{output_path}")

    def export_items_to_excel(self, items, default_name=None, parent=None):
        if not items:
            messagebox.showwarning("다운로드 대상 없음", "다운로드할 항목을 먼저 선택하거나 입력하세요.", parent=parent)
            return False
        template_path = self.get_resource_path("견적용.xlsx")
        if not os.path.exists(template_path):
            messagebox.showerror("파일 오류", f"'{template_path}' 파일이 존재하지 않습니다.", parent=parent)
            return False
        output_path = filedialog.asksaveasfilename(parent=parent or self.root, title="다운로드 저장 위치 선택", defaultextension=".xlsx", initialfile=(default_name or f"선택견적_{datetime.now():%Y-%m-%d}.xlsx"), filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not output_path:
            return False
        wb = openpyxl.load_workbook(template_path)
        if "기계" not in wb.sheetnames:
            messagebox.showerror("시트 오류", "'기계' 시트를 찾을 수 없습니다.", parent=parent)
            return False
        ws = wb["기계"]
        saved_count = self._save_items_to_workbook(ws, items, False)
        wb.save(output_path)
        messagebox.showinfo("선택 다운로드 완료", f"{saved_count}개 항목을 다운로드용 견적 파일로 저장했습니다.\n\n{output_path}", parent=parent)
        return True

    def export_selected_items(self):
        selected_items = [item for item in self.data if self.has_item_data(item) and item["no"] in self.selected_nos]
        self.export_items_to_excel(selected_items, default_name=f"선택견적_{datetime.now():%Y-%m-%d}.xlsx", parent=self.root)


if __name__ == "__main__":
    root = tk.Tk()
    app = MachineEstimateApp(root)
    root.mainloop()
