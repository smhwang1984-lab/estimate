"""견적 양식(.xlsx) 읽기·쓰기.

v0.0.7부터 열 위치를 코드에 고정하지 않는다. `resolve_columns()`가 6행 헤더 문구를 읽어
이 파일이 실제로 어떤 배치인지 그때그때 판단한다. 예전에 쓰던 다른 배치의 양식
(기종 열이 없거나, 치구/프로그램이 한 칸으로 합쳐진 구버전 등)을 업로드해도
엉뚱한 열에 값이 들어가지 않게 하기 위해서다.

행이 모자라면 openpyxl의 insert_rows를 쓰지 않는다. 병합 범위가 따라가지 않는 것을
직접 재현해 확인했다(8건째부터 기계 시트 푸터가 밀리고 견적서가 #REF!가 나던 원인).
대신 '합계 행부터 그 아래(푸터 포함)를 통째로 아래로 옮기고, 비게 된 자리에
서식을 갖춘 새 데이터 행을 채우는' 방식(shift_block_down)을 쓴다.

    5행: 공정별 단가   6행: 헤더   7행부터: 입력 행   헤더 열의 '합계' 문구가 있는 행 = 합계 행

v0.1.0: 출력 파일에 수식을 남기지 않는다(요청 1-3). 금액·순번·합계를 전부 계산된
값으로 적는다 -- 화면과 같은 계산 결과가 나가도록 core/pricing.calc_row()를 그대로 쓴다.
엑셀에서 단가·시간을 직접 고쳐도 더 이상 재계산되지 않는다는 뜻이니, 그 값이 필요하면
프로그램에서 고쳐 다시 출력해야 한다. 같은 김에 견적서 시트의 숨은 항목 행(요청 1-1)과
Comment 칸의 행 높이(요청 1-2)도 여기서 맞춘다.
"""

import os
import re
from copy import copy
from datetime import datetime

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, range_boundaries

from . import paths, settings as settings_store
from .config import MACHINE_KEYS
from .model import create_blank_item, material_text, safe_float, safe_int
from .pricing import calc_row

SHEET_NAME = "기계"
ESTIMATE_SHEET_NAME = "견적서"
TEMPLATE_NAME = "견적용.xlsx"

HEADER_ROW = 6
FIRST_DATA_ROW = 7
RATE_ROW = 5
ES_FIRST_ROW = 15
ES_ROW_OFFSET = ES_FIRST_ROW - FIRST_DATA_ROW  # 견적서 행 - 기계 행
ES_COLS = {"no": 2, "part_no": 3, "part_name": 5, "final_price": 7}  # 견적서 자체 열은 항상 고정

# 기계 시트 푸터(회사명 / 날짜 / 작성자)는 합계 행 바로 아래에 붙어 있고, 데이터 행이 늘면
# 합계 행과 함께 통째로 밀려 내려간다(직접 확인: 7건 Q17~19, 20건 Q30~32, 45건 Q55~57).
# 그래서 절대 행 번호가 아니라 합계 행 기준 상대 위치로 찾는다.
FOOTER_COL = 17                  # Q열
FOOTER_COMPANY_OFFSET = 3        # 합계행 +3: 회사명
FOOTER_DATE_OFFSET = 4           # 합계행 +4: 발행 날짜(yyyy-mm-dd)
FOOTER_WRITER_OFFSET = 5         # 합계행 +5: 직위 + 작성자

# v0.0.9: 단가/최종단가 열이 좁아 금액이 `####`으로 나오던 것을 넓힌다(요청 3-3).
# 회계 서식(`_-* #,##0_-;...`)은 숫자 좌우에 여백까지 잡아먹어 기본 폭으로는 백만 단위부터 잘린다.
# 서식 자체는 손대지 않는다 -- 보이게만 하면 되는 요청이다.
PRICE_COL_WIDTH = 15

# 견적서 시트 왼쪽 위(받는 업체) / 오른쪽 위(우리 쪽 연락처) 칸 위치와 문구 틀.
# 설정창에는 알맹이만 입력받고, 양식 문구는 여기서 조립한다.
ES_CLIENT_CELLS = [
    ("B4", "{value} 귀중", "name"),
    ("B5", "  담당자 : {value}", "manager"),
    ("B6", "  부   서 : {value}", "dept"),
    ("B8", "  사업명 : {value}", "project"),
]
ES_SUPPLIER_CELLS = [
    ("F5", "대표자 : {value}    (인)", "ceo"),
    ("F6", "사업자등록번호 : {value}", "biz_no"),
    ("F7", "주소 : {value}", "address"),
    ("F8", "전화번호 : {value}", "phone"),
    ("F9", "담당자 : {value}", "contact"),
]
ES_DATE_CELL = "B7"  # 기계 시트 푸터의 날짜를 참조한다(=기계!Q##).

# Material 칸에 붙여 둔 경도 표기를 되읽기 위한 규칙. `HRC58~62`, `HRC 58 ~ 62`, `HRC58` 모두 받는다.
HRC_PATTERN = re.compile(r"\s*HRC\s*([\d.]+)?\s*(?:~|-)?\s*([\d.]+)?\s*$", re.IGNORECASE)

# 헤더 문구가 살짝 달라도(구버전 양식 등) 찾아내기 위한 부분 일치 키워드.
# 앞쪽 항목이 우선한다 -- '프로그램 및치구'처럼 한 칸에 합쳐진 헤더는 먼저 매칭되는
# 키가 그 칸을 차지하고, 나머지 하나는 이 파일에서 열이 없는 것으로 처리한다
# (두 키에 같은 값을 중복으로 넣으면 금액이 두 배로 계산되므로).
MACHINE_KEYWORDS = [
    ("m_prog", ("프로그램",)),
    ("m_jig", ("치구",)),
    ("m_5axis", ("5축",)),
    ("m_4axis", ("4축",)),
    ("m_3axis", ("3축",)),
    ("m_lathe", ("선반",)),
    ("m_general", ("범용",)),
    ("m_finish", ("사상",)),
    ("m_cmm", ("CMM", "3차원")),
    ("m_grind", ("연삭", "연마", "와이어", "EDM")),
]


def _openpyxl():
    """openpyxl은 불러오는 데만 3초 가까이 걸린다.

    프로그램을 켤 때가 아니라 엑셀을 실제로 읽고 쓸 때 처음 불러오도록 미뤄서
    시작 시간을 줄인다(요청: "프로그램이 무거워").
    """
    import openpyxl
    return openpyxl


class TemplateNotFound(Exception):
    pass


class SheetNotFound(Exception):
    pass


def get_template_path():
    path = paths.get_resource_path(TEMPLATE_NAME)
    if not os.path.exists(path):
        raise TemplateNotFound(path)
    return path


def _get_sheet(wb, name):
    if name not in wb.sheetnames:
        raise SheetNotFound(name)
    return wb[name]


# ---------- 열 위치 판단 ----------

def resolve_columns(ws):
    """6행 헤더 문구를 읽어 이 파일의 실제 열 배치를 돌려준다."""
    texts = {}
    for col in range(1, 41):
        value = ws.cell(row=HEADER_ROW, column=col).value
        if value not in (None, ""):
            texts[col] = str(value).strip()

    def find_exact(*labels):
        for col in sorted(texts):
            if texts[col] in labels:
                return col
        return None

    used = set()

    def find_contains(*keywords):
        for col in sorted(texts):
            if col in used:
                continue
            if any(kw in texts[col] for kw in keywords):
                used.add(col)
                return col
        return None

    machine = {}
    for key, keywords in MACHINE_KEYWORDS:
        machine[key] = find_contains(*keywords)

    return {
        "model": find_exact("기종"),
        "part_no": find_exact("품번"),
        "part_name": find_exact("품명"),
        "comment": find_exact("Coment", "Comment"),
        "possible": find_exact("가능여부"),
        "qty": find_exact("Qty"),
        "material": find_exact("Material"),
        "size": find_exact("Size"),
        "unit_price": find_exact("단가"),
        "final_price": find_exact("최종단가"),
        "sum": find_exact("SUM"),
        "machine": machine,
    }


def _max_used_col(cols):
    values = [v for k, v in cols.items() if k != "machine" and v]
    values += [v for v in cols["machine"].values() if v]
    return max(values) if values else 24


def _merge_span(ws, row, col):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.max_col - rng.min_col + 1
    return 1


def _print_area_bounds(ws):
    """현재 인쇄 영역의 (좌열, 상행, 우열, 하행). 없으면 None."""
    if not ws.print_area:
        return None
    ref = str(ws.print_area)
    if "!" in ref:
        ref = ref.split("!", 1)[1]
    ref = ref.replace("$", "")
    try:
        return range_boundaries(ref)
    except ValueError:
        return None


def _set_print_area(ws, min_col, min_row, max_col, max_row):
    ws.print_area = (f"${get_column_letter(min_col)}${min_row}:"
                     f"${get_column_letter(max_col)}${max_row}")


# ---------- 블록 이동 (insert_rows 대신 씀. 병합도 함께 옮긴다) ----------

def shift_block_down(ws, from_row, delta, max_col):
    """from_row 이상의 모든 내용(값·서식·행높이·병합)을 delta만큼 아래로 옮긴다."""
    if delta <= 0:
        return
    last_row = max(ws.max_row, from_row)
    merges_to_move = [rng for rng in list(ws.merged_cells.ranges) if rng.min_row >= from_row]
    for rng in merges_to_move:
        ws.unmerge_cells(str(rng))

    for row in range(last_row, from_row - 1, -1):
        target = row + delta
        for col in range(1, max_col + 1):
            src = ws.cell(row=row, column=col)
            dst = ws.cell(row=target, column=col)
            dst.value = src.value
            if src.has_style:
                dst._style = copy(src._style)
            dst.number_format = src.number_format
        if row in ws.row_dimensions:
            ws.row_dimensions[target].height = ws.row_dimensions[row].height

    for row in range(from_row, from_row + delta):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    for rng in merges_to_move:
        ws.merge_cells(start_row=rng.min_row + delta, start_column=rng.min_col,
                       end_row=rng.max_row + delta, end_column=rng.max_col)


def copy_row_style(ws, source_row, target_row, max_col):
    for col in range(1, max_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def prepare_blank_row(ws, row, cols):
    """항목이 들어가지 않는 칸의 순번(A)·SUM(V)·단가(W)·최종단가(X)를 비운다.

    양식 원본 7~14행에는 처음부터 수식이 박혀 있어(`=ROW()-6` 등), 항목을 채우지 않은
    칸도 그대로 두면 0원이 찍힌 채로 남는다. 수식을 없애면서(요청 1-3) 함께 정리한다.
    """
    ws.cell(row=row, column=1).value = None
    for key in ("sum", "unit_price", "final_price"):
        col = cols.get(key)
        if col:
            ws.cell(row=row, column=col).value = None


def write_row_values(ws, row, item, rates, cols):
    """항목 한 건의 순번·시간합계·단가합계·최종단가를 계산된 값으로 적는다(요청 1-3).

    화면과 엑셀이 같은 금액을 보이도록 core/pricing.calc_row()로 계산한다 -- 계산 규칙을
    한 곳에만 두는 지금 규칙(pricing.py 첫머리 참고)을 그대로 따른다.
    """
    ws.cell(row=row, column=1, value=row - HEADER_ROW)  # 기존 '=ROW()-6'과 같은 값
    sum_time, unit_price, final_price = calc_row(item, rates)
    if cols["sum"]:
        ws.cell(row=row, column=cols["sum"], value=round(sum_time, 2))
    if cols["unit_price"]:
        ws.cell(row=row, column=cols["unit_price"], value=round(unit_price))
    if cols["final_price"]:
        ws.cell(row=row, column=cols["final_price"], value=final_price)
    return final_price


ROW_LINE_HEIGHT = 16.0   # 한 줄당 높이(pt). 사용자가 손으로 맞춰 둔 E29 사례(7줄 -> 112)로 검증한 값.
DEFAULT_ROW_HEIGHT = 30.0  # 양식 기본 데이터 행 높이.


def _wrapped_line_count(text, col_width_chars):
    """줄바꿈(\\n)과, 한 줄이 열 폭을 넘길 때의 자동 줄바꿈을 합쳐 필요한 줄 수를 센다.

    엑셀의 실제 글자 폭 계산과는 다른 근사치다(고정폭으로 가정). 정확한 렌더링은
    엑셀에서 직접 열어야 확인할 수 있다.
    """
    if text in (None, ""):
        return 1
    chars_per_line = max(1, int(col_width_chars))
    total = 0
    for paragraph in str(text).split("\n"):
        if not paragraph:
            total += 1
            continue
        total += -(-len(paragraph) // chars_per_line)  # 올림 나눗셈
    return max(1, total)


def fit_row_height(ws, row, max_col):
    """줄바꿈(wrap_text)이 켜진 칸의 내용이 몇 줄이 되는지 세어 행 높이를 맞춘다(요청 1-2).

    wrap_text가 꺼진 칸(품명 등)은 셀 안에서 줄바꿈이 되지 않으니(그대로 잘리거나 옆 칸을
    침범한다) 높이를 늘려도 의미가 없어 계산에서 뺀다 -- 지금 양식은 Comment(E열)만
    wrap_text가 켜져 있다(확인함).
    """
    max_lines = 1
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if not cell.alignment or not cell.alignment.wrap_text or cell.value in (None, ""):
            continue
        width = ws.column_dimensions[get_column_letter(col)].width or 10
        max_lines = max(max_lines, _wrapped_line_count(cell.value, width))
    # customHeight는 openpyxl에서 읽기 전용 계산 속성이다(height가 있으면 자동으로 True).
    ws.row_dimensions[row].height = max(DEFAULT_ROW_HEIGHT, max_lines * ROW_LINE_HEIGHT)


def merge_size_cell(ws, row, cols):
    """Size 헤더가 여러 칸(I6:K6 등) 병합돼 있으면 데이터 행도 같은 폭으로 병합하고
    가운데 정렬한다. 지금은 값이 I열 한 칸에만 들어가 사이즈 하나가 3열로 쪼개져
    보이는 문제(요청 1번)를 고친다. 헤더가 1칸뿐인 구버전 양식은 손대지 않는다.
    """
    if not cols["size"]:
        return
    span = _merge_span(ws, HEADER_ROW, cols["size"])
    if span <= 1:
        return
    start_col = cols["size"]
    end_col = start_col + span - 1
    # 다시 저장하는 경우 이미 병합돼 있을 수 있다 -- 겹치는 병합을 정리하고 새로 맞춘다
    # (openpyxl은 이미 병합된 범위를 다시 merge_cells 하면 오류를 낸다).
    for rng in [r for r in list(ws.merged_cells.ranges)
               if r.min_row == row and r.max_row == row
               and r.min_col >= start_col and r.max_col <= end_col]:
        ws.unmerge_cells(str(rng))
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    ws.cell(row=row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")


def row_has_input_data(ws, row, cols):
    check_cols = [cols[key] for key in
                  ("part_no", "part_name", "comment", "possible", "qty", "material", "size")
                  if cols[key]]
    check_cols += [col for col in cols["machine"].values() if col]
    for col in check_cols:
        if ws.cell(row=row, column=col).value not in (None, ""):
            return True
    return False


def find_summary_row(ws, cols):
    """'합계' 문구가 있는 행을 헤더 문구가 아니라 실제로 찾는다(행이 몇 개든 안전)."""
    label_col = cols.get("unit_price") or cols.get("final_price")
    if not label_col:
        return FIRST_DATA_ROW
    row = FIRST_DATA_ROW
    limit = FIRST_DATA_ROW + 2000
    while row < limit:
        value = ws.cell(row=row, column=label_col).value
        if isinstance(value, str) and "합계" in value.replace(" ", ""):
            return row
        row += 1
    return FIRST_DATA_ROW


def write_summary_label(ws, cols, summary_row):
    """합계 행의 '합계' 문구만 적는다. 금액은 항목을 다 쓴 뒤 write_summary_total()이 채운다."""
    if cols["unit_price"]:
        ws.cell(row=summary_row, column=cols["unit_price"], value="합계")


def write_summary_total(ws, cols, summary_row, total):
    """항목별 최종단가를 다 더한 값을 합계 행에 적는다(요청 1-3, 수식 대신 값)."""
    if cols["final_price"]:
        ws.cell(row=summary_row, column=cols["final_price"], value=total)


def layout_machine_sheet(ws, cols, need_rows):
    """기계 시트에 need_rows개의 데이터 행이 들어갈 자리를 만들고 합계 행을 돌려준다.

    모자라면 합계 행(및 그 아래 푸터)을 통째로 아래로 밀고, 비게 된 자리에
    7행 서식을 복사한 새 데이터 행을 채운다. 인쇄 영역도 늘어난 만큼 넓힌다.
    """
    max_col = _max_used_col(cols)
    summary_row = find_summary_row(ws, cols)
    capacity = summary_row - FIRST_DATA_ROW
    if need_rows > capacity:
        additional = need_rows - capacity
        bounds = _print_area_bounds(ws)
        shift_block_down(ws, summary_row, additional, max_col)
        # shift_block_down이 이미 값(A/V/W/X 포함)을 비워 뒀으므로 여기서 수식을 새로
        # 넣지 않는다(v0.1.0, 요청 1-3) -- 서식·병합만 갖추고 값은 write_row_values()가 채운다.
        for row in range(summary_row, summary_row + additional):
            copy_row_style(ws, FIRST_DATA_ROW, row, max_col)
            merge_size_cell(ws, row, cols)
        summary_row += additional
        if bounds:
            min_col, min_row, max_col_b, max_row_b = bounds
            _set_print_area(ws, min_col, min_row, max_col_b, max_row_b + additional)
    write_summary_label(ws, cols, summary_row)
    return summary_row


def sync_estimate_sheet(es, gigye_ws, gigye_cols, gigye_summary_row):
    """기계 시트의 실제 데이터 행 수에 맞춰 견적서 시트의 항목 칸 수·값을 맞추고,
    항목이 든 행만 보이도록 숨김을 정리한다(요청 1-1, 1-3).

    v0.1.0부터 참조 수식(`=기계!$C7`) 대신 기계 시트에 이미 값으로 적힌 품번·품명·
    최종단가를 그대로 복사해 적는다 -- write_row_values()가 그 값들을 먼저 써 두므로
    여기서는 읽기만 하면 된다.
    """
    item_count = gigye_summary_row - FIRST_DATA_ROW

    summary_row = ES_FIRST_ROW
    limit = ES_FIRST_ROW + 2000
    while summary_row < limit:
        value = es.cell(row=summary_row, column=ES_COLS["no"]).value
        if isinstance(value, str) and "합계" in value.replace(" ", ""):
            break
        summary_row += 1
    else:
        return  # 예상 밖 구조라 손대지 않는다.

    capacity = summary_row - ES_FIRST_ROW
    if item_count > capacity:
        additional = item_count - capacity
        bounds = _print_area_bounds(es)
        shift_block_down(es, summary_row, additional, 8)
        for row in range(summary_row, summary_row + additional):
            copy_row_style(es, ES_FIRST_ROW, row, 8)
            for rng in (f"C{row}:D{row}", f"E{row}:F{row}", f"G{row}:H{row}"):
                es.merge_cells(rng)
        summary_row += additional
        if bounds:
            min_col, min_row, max_col_b, max_row_b = bounds
            _set_print_area(es, min_col, min_row, max_col_b, max_row_b + additional)

    # 기계 시트에서 실제로 값이 있는 행만 옮겨 적는다. 예전에는 "< gigye_summary_row"
    # (양식의 칸 수)만으로 판단해서, 항목보다 빈 칸이 많이 남으면 그 빈 칸을 그대로 참조해
    # 견적서에 품번/품명 '0', 단가 '₩0'인 쓰레기 행이 찍혔다(재현 완료).
    #
    # row_has_input_data만으로 바꿔서도 안 된다 -- 기계 시트 푸터(17~19행)가 데이터
    # 열과 같은 열 위치를 쓴다(예: '사상' 열 Q=17번째 열에 푸터 라벨 '㈜텍스타'가 Q17에
    # 앉아 있다, 'coment :' 라벨은 D18에 있고 D는 품명 열이다). 상한 없이 검사하면 이
    # 푸터 라벨을 항목 데이터로 착각한다(직접 재현 확인). 그래서 "기계 시트의 실제 데이터
    # 구간(FIRST_DATA_ROW ~ gigye_summary_row 미만) 안에 있고, 그 행에 실제 값이 있을 때"
    # 둘 다를 만족해야 옮겨 적는다.
    part_no_col = gigye_cols["part_no"]
    part_name_col = gigye_cols["part_name"]
    final_price_col = gigye_cols["final_price"]
    total = 0
    filled_count = 0
    for row in range(ES_FIRST_ROW, summary_row):
        src_row = row - ES_ROW_OFFSET
        has_data = (FIRST_DATA_ROW <= src_row < gigye_summary_row
                   and row_has_input_data(gigye_ws, src_row, gigye_cols))
        # 주의: ws.cell(..., value=None)은 openpyxl에서 "값을 지운다"가 아니라
        # "value 인자를 안 준 것"으로 취급되어 기존 값이 그대로 남는다(직접 재현 확인).
        # 반드시 .value 속성에 직접 대입해야 실제로 비워진다.
        if part_no_col:
            es.cell(row=row, column=ES_COLS["part_no"]).value = (
                str(gigye_ws.cell(row=src_row, column=part_no_col).value or "") if has_data else None)
        if part_name_col:
            es.cell(row=row, column=ES_COLS["part_name"]).value = (
                str(gigye_ws.cell(row=src_row, column=part_name_col).value or "") if has_data else None)
        price_value = (gigye_ws.cell(row=src_row, column=final_price_col).value
                       if (has_data and final_price_col) else None)
        if final_price_col:
            es.cell(row=row, column=ES_COLS["final_price"]).value = price_value
        if has_data:
            filled_count += 1
            total += price_value or 0

    # 요청 1-1: 실제 항목이 든 행만 보이고, 그 다음 한 행은 비운 채 보이고(구분선 대신),
    # 나머지는 숨긴다. 양식 원본은 22행부터 끝까지 hidden이라 항목이 8건만 넘어도
    # 견적서 화면·인쇄에서 사라졌었다(재현 확인).
    blank_row = ES_FIRST_ROW + filled_count
    for row in range(ES_FIRST_ROW, summary_row):
        if row < blank_row:
            es.row_dimensions[row].hidden = False
            es.cell(row=row, column=ES_COLS["no"]).value = row - 14
        elif row == blank_row:
            es.row_dimensions[row].hidden = False
            es.cell(row=row, column=ES_COLS["no"]).value = None
        else:
            es.row_dimensions[row].hidden = True
            es.cell(row=row, column=ES_COLS["no"]).value = None
    es.row_dimensions[summary_row].hidden = False

    if final_price_col:
        es.cell(row=summary_row, column=ES_COLS["final_price"], value=total)


# ---------- 읽기 ----------

def split_material(text):
    """Material 칸 문자열을 (소재, 열처리여부, HRC Min, HRC Max)로 나눈다.

    저장할 때 붙인 `HRC58~62`를 그대로 다시 카드로 되돌리기 위한 것이다.
    경도 표기가 없으면 소재만 돌려주고 열처리는 꺼진 상태가 된다.
    """
    text = str(text or "").strip()
    if not text:
        return "", False, "", ""
    match = HRC_PATTERN.search(text)
    if not match:
        return text, False, "", ""
    low, high = match.group(1) or "", match.group(2) or ""
    if not low and not high:
        # `HRC`만 적혀 있는 경우. 열처리는 켜 두되 값은 비운다.
        return text[:match.start()].strip(), True, "", ""
    if not high:
        # `HRC58` 한쪽만 적힌 경우는 Min으로 본다.
        return text[:match.start()].strip(), True, low, ""
    return text[:match.start()].strip(), True, low, high


def read_cards_from_workbook(file_path):
    """업로드한 기계 시트에서 입력된 행만 카드로 읽어 온다."""
    wb = _openpyxl().load_workbook(file_path, data_only=False)
    ws = _get_sheet(wb, SHEET_NAME)
    cols = resolve_columns(ws)
    summary_row = find_summary_row(ws, cols)
    abs_path = os.path.abspath(file_path)
    created_at = datetime.fromtimestamp(os.path.getmtime(abs_path)).strftime("%Y-%m-%d")
    source_month = os.path.basename(os.path.dirname(abs_path))
    size_span = _merge_span(ws, HEADER_ROW, cols["size"]) if cols["size"] else 0

    cards = []
    for row in range(FIRST_DATA_ROW, summary_row):
        if not row_has_input_data(ws, row, cols):
            continue
        item = create_blank_item(len(cards) + 1)
        item["created_at"] = created_at
        item["source_file"] = os.path.basename(abs_path)
        item["source_month"] = source_month
        item["save_pending"] = True
        item["model"] = str(ws.cell(row=row, column=cols["model"]).value or "") if cols["model"] else ""
        item["part_no"] = str(ws.cell(row=row, column=cols["part_no"]).value or "") if cols["part_no"] else ""
        item["part_name"] = (str(ws.cell(row=row, column=cols["part_name"]).value or "")
                             if cols["part_name"] else "")
        item["comment"] = str(ws.cell(row=row, column=cols["comment"]).value or "") if cols["comment"] else ""
        item["possible"] = (str(ws.cell(row=row, column=cols["possible"]).value or "가능")
                            if cols["possible"] else "가능")
        item["qty"] = safe_int(ws.cell(row=row, column=cols["qty"]).value, 1) if cols["qty"] else 1
        raw_material = (str(ws.cell(row=row, column=cols["material"]).value or "")
                        if cols["material"] else "")
        item["material"], item["heat_treat"], item["hrc_min"], item["hrc_max"] = split_material(raw_material)
        if cols["size"] and size_span:
            size_parts = [ws.cell(row=row, column=cols["size"] + i).value for i in range(size_span)]
            item["size"] = " x ".join(str(v) for v in size_parts if v not in (None, ""))
        for key, col in cols["machine"].items():
            item[key] = safe_float(ws.cell(row=row, column=col).value) if col else 0.0
        cards.append(item)
    return cards


# ---------- 쓰기 ----------

def _write_item_values(ws, row, item, cols):
    # 주의: ws.cell(..., value=None)은 openpyxl에서 "값을 지운다"가 아니라
    # "value 인자를 안 준 것"으로 취급되어 기존 값이 그대로 남는다(직접 재현 확인).
    # 사용자가 필드를 비우고 다시 저장했을 때 이전 값이 남지 않도록 .value에 직접 대입한다.
    def set_cell(col, value):
        ws.cell(row=row, column=col).value = value

    def text_or_none(value):
        value = value.strip() if value else ""
        return value or None

    if cols["model"]:
        set_cell(cols["model"], text_or_none(item["model"]))
    if cols["part_no"]:
        set_cell(cols["part_no"], text_or_none(item["part_no"]))
    if cols["part_name"]:
        set_cell(cols["part_name"], text_or_none(item["part_name"]))
    if cols["comment"]:
        set_cell(cols["comment"], text_or_none(item["comment"]))
    if cols["possible"]:
        set_cell(cols["possible"], item["possible"])
    if cols["qty"]:
        set_cell(cols["qty"], item["qty"])
    if cols["material"]:
        # 열처리를 체크한 항목은 `SUS304 HRC58~62`처럼 경도를 붙여 적는다(요청 5-2, 5-3).
        # 기계 시트에 열처리 전용 열이 없어 Material 칸에 함께 담는다.
        set_cell(cols["material"], text_or_none(material_text(item)))
    if cols["size"]:
        set_cell(cols["size"], text_or_none(item["size"]))
    for key, col in cols["machine"].items():
        if not col:
            continue
        value = item[key]
        set_cell(col, value if value > 0 else None)


def write_items_to_sheet(ws, items, rates):
    """빈 양식의 기계 시트에 카드 목록을 위에서부터 차례로 채워 넣는다.

    v0.0.9에서 날짜별 누적 저장을 없애면서(요청 4-2) 출력은 항상 빈 양식에서 시작한다.
    그래서 예전처럼 "이 카드가 지난번에 몇 행에 들어갔는지" 기억했다가 그 자리를 다시 쓰는
    로직은 필요 없어졌다 -- 카드 순서 그대로 위에서부터 채운다.
    """
    cols = resolve_columns(ws)

    for key, col in cols["machine"].items():
        if col:
            ws.cell(row=RATE_ROW, column=col, value=rates[key])

    summary_row = find_summary_row(ws, cols)
    free_rows = [row for row in range(FIRST_DATA_ROW, summary_row)
                if not row_has_input_data(ws, row, cols)]

    need_new = max(0, len(items) - len(free_rows))
    if need_new:
        capacity = summary_row - FIRST_DATA_ROW
        summary_row = layout_machine_sheet(ws, cols, capacity + need_new)
        free_rows = [row for row in range(FIRST_DATA_ROW, summary_row)
                    if not row_has_input_data(ws, row, cols)]
    else:
        write_summary_label(ws, cols, summary_row)

    max_col = _max_used_col(cols)
    saved_count = 0
    total = 0
    for item in items:
        row = free_rows.pop(0)
        copy_row_style(ws, FIRST_DATA_ROW, row, max_col)
        merge_size_cell(ws, row, cols)
        _write_item_values(ws, row, item, cols)
        total += write_row_values(ws, row, item, rates, cols)
        fit_row_height(ws, row, max_col)
        saved_count += 1
    # 다 채우고 남은 칸(양식 원본에 처음부터 있던 수식 포함)은 비워 둔다(요청 1-3).
    for row in free_rows:
        prepare_blank_row(ws, row, cols)
    write_summary_total(ws, cols, summary_row, total)
    widen_price_columns(ws, cols)
    return saved_count, summary_row, cols


def _sync_estimate_if_present(wb, gigye_ws, gigye_cols, gigye_summary_row):
    if ESTIMATE_SHEET_NAME in wb.sheetnames:
        sync_estimate_sheet(wb[ESTIMATE_SHEET_NAME], gigye_ws, gigye_cols, gigye_summary_row)


# ---------- 설정값 반영 (회사/작성자/견적서 상단/날짜) ----------

def write_machine_footer(ws, summary_row, config_values, today=None):
    """기계 시트 푸터에 회사명·발행 날짜·작성자를 적고, 적어 넣은 날짜 문자열을 돌려준다.

    푸터는 합계 행과 함께 밀려 내려가므로 위치를 상대 계산한다(FOOTER_*_OFFSET 참고).
    """
    company = config_values["company"]
    date_row = summary_row + FOOTER_DATE_OFFSET
    today = today or datetime.now()
    ws.cell(row=summary_row + FOOTER_COMPANY_OFFSET, column=FOOTER_COL).value = (
        str(company.get("name", "")).strip() or None)
    # 요청 3-6: 양식에 'YYYY-MM-DD'라고만 적혀 있던 자리에 실제 출력 날짜를 넣는다.
    date_text = today.strftime("%Y-%m-%d")
    ws.cell(row=date_row, column=FOOTER_COL).value = date_text
    ws.cell(row=summary_row + FOOTER_WRITER_OFFSET, column=FOOTER_COL).value = (
        settings_store.writer_line(company) or None)
    return date_text


def write_estimate_header(es, config_values, date_text):
    """견적서 시트 맨 위(받는 업체 / 우리 쪽 연락처 / 날짜)를 설정값으로 채운다(요청 7-1).

    양식에는 `회사명㈜ 귀중`, `사업명` 같은 플레이스홀더가 그대로 남아 있었다.
    설정창에는 알맹이만 입력받고 `담당자 :` 같은 양식 문구는 여기서 붙인다.
    """
    def fill(cells, values):
        for address, template, key in cells:
            value = str(values.get(key, "")).strip()
            es[address] = template.format(value=value) if value else None

    fill(ES_CLIENT_CELLS, config_values["client"])
    fill(ES_SUPPLIER_CELLS, config_values["supplier"])
    # v0.1.0: 기계 시트 푸터를 가리키던 참조 수식(`=기계!Q##`) 대신 같은 날짜 문자열을
    # 그대로 적는다(요청 1-3) -- 항목 수에 따라 푸터가 밀려도 다시 계산할 필요가 없다.
    es[ES_DATE_CELL] = date_text


def widen_price_columns(ws, cols):
    """단가·최종단가 열이 좁아 금액이 `####`으로 나오던 것을 넓힌다(요청 3-3)."""
    for key in ("unit_price", "final_price"):
        col = cols.get(key)
        if not col:
            continue
        letter = get_column_letter(col)
        current = ws.column_dimensions[letter].width or 0
        if current < PRICE_COL_WIDTH:
            ws.column_dimensions[letter].width = PRICE_COL_WIDTH


def export_items(items, rates, output_path, config_values=None):
    """빈 양식에 선택한 카드만 담아 별도 파일로 저장한다.

    v0.0.9부터 저장할 때마다 설정창 값(회사명·작성자·견적서 상단)과 오늘 날짜를 함께 적는다.
    """
    config_values = config_values or settings_store.load()
    template_path = get_template_path()
    wb = _openpyxl().load_workbook(template_path)
    ws = _get_sheet(wb, SHEET_NAME)
    saved_count, summary_row, cols = write_items_to_sheet(ws, items, rates)
    _sync_estimate_if_present(wb, ws, cols, summary_row)
    date_text = write_machine_footer(ws, summary_row, config_values)
    if ESTIMATE_SHEET_NAME in wb.sheetnames:
        write_estimate_header(wb[ESTIMATE_SHEET_NAME], config_values, date_text)
    wb.save(output_path)
    return saved_count
