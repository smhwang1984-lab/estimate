import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
import re

class MachineEstimateApp:
    def __init__(self, root):
        self.root = root
        self.root.title("기계 시트 표준 견적 입력 시스템")
        self.root.geometry("1200x650")

        self.excel_file = "견적용.xlsx"
        self.data = []

        self.rates = {
            "m_5axis": 70000, "m_4axis": 50000, "m_3axis": 40000, "m_lathe": 35000,
            "m_general": 20000, "m_finish": 15600, "m_cmm": 30000, "m_grind": 35000,
            "m_jig": 35000, "m_prog": 35000
        }

        self.mach_keys = [
            ("m_5axis", "5축 NC (7만/h)"), ("m_4axis", "4축 (5만/h)"),
            ("m_3axis", "3축 NC (4만/h)"), ("m_lathe", "NC 선반 (3.5만/h)"),
            ("m_general", "범용 (2만/h)"), ("m_finish", "사상 (1.56만/h)"),
            ("m_cmm", "CMM (3만/h)"), ("m_grind", "연삭/와이어 (3.5만/h)"),
            ("m_jig", "치구 (3.5만/h)"), ("m_prog", "프로그램 (3.5만/h)")
        ]

        # 상단 버튼 및 안내
        btn_frame = ttk.Frame(root, padding=10)
        btn_frame.pack(fill=tk.X)

        ttk.Button(btn_frame, text="📂 '견적용.xlsx' 불러오기", command=self.load_from_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="💾 '견적용.xlsx' 저장하기", command=self.save_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="➕ 행 추가", command=self.add_row).pack(side=tk.LEFT, padx=5)

        guide_label = ttk.Label(btn_frame, text="💡 [NO] 또는 행을 더블클릭하면 상세 입력 팝업창이 뜹니다.", font=("맑은 고딕", 10, "bold"))
        guide_label.pack(side=tk.RIGHT, padx=10)

        # Treeview (기계 시트 열 구조 반영)
        self.columns = ("no", "part_no", "part_name", "comment", "possible", "qty", "material", "size",
                        "sum_time", "unit_price", "final_price")

        self.tree = ttk.Treeview(root, columns=self.columns, show="headings", height=20)

        headers = {
            "no": "NO", "part_no": "품번", "part_name": "품명", "comment": "Comment",
            "possible": "가능여부", "qty": "Qty", "material": "Material", "size": "Size",
            "sum_time": "SUM(시간)", "unit_price": "단가(원)", "final_price": "최종단가(절사)"
        }

        widths = {"no": 45, "part_no": 120, "part_name": 140, "comment": 110, "possible": 70,
                  "qty": 55, "material": 110, "size": 140, "sum_time": 90, "unit_price": 110, "final_price": 110}

        for col in self.columns:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], anchor="center")

        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.tree.bind("<Double-1>", self.on_double_click)

        # 프로그램 실행 시 데이터 읽기 시도 (없을 경우 빈 데이터 10행 생성)
        if not self.load_from_excel(show_message=False):
            self.init_empty_data()
            self.refresh_table()

    def init_empty_data(self):
        self.data = []
        for i in range(1, 11):
            self.data.append({
                "no": i, "part_no": "", "part_name": "", "comment": "", "possible": "가능",
                "qty": 1, "material": "", "size": "",
                "m_5axis": 0.0, "m_4axis": 0.0, "m_3axis": 0.0, "m_lathe": 0.0,
                "m_general": 0.0, "m_finish": 0.0, "m_cmm": 0.0, "m_grind": 0.0,
                "m_jig": 0.0, "m_prog": 0.0
            })

    def calc_row(self, item):
        times = [item[k[0]] for k in self.mach_keys]
        sum_time = sum(times)

        cost_sum = sum(item[k[0]] * self.rates[k[0]] for k in self.mach_keys)
        unit_price = cost_sum * item["qty"]
        final_price = int(unit_price // 1000 * 1000)
        return sum_time, unit_price, final_price

    def refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for item in self.data:
            sum_time, unit_price, final_price = self.calc_row(item)
            has_data = bool(item["part_no"] or item["part_name"])

            self.tree.insert("", tk.END, iid=str(item["no"]), values=(
                item["no"],
                item["part_no"] if item["part_no"] else "-",
                item["part_name"] if item["part_name"] else "-",
                item["comment"] if item["comment"] else "-",
                item["possible"],
                item["qty"],
                item["material"] if item["material"] else "-",
                item["size"] if item["size"] else "-",
                f"{sum_time:.1f}" if has_data else "0",
                f"{int(unit_price):,}" if has_data else "0",
                f"{int(final_price):,}" if has_data else "0"
            ))

    def on_double_click(self, event):
        selected = self.tree.selection()
        if selected:
            no = int(selected[0])
            self.open_popup(no)

    def add_row(self):
        new_no = max([item["no"] for item in self.data], default=0) + 1
        self.data.append({
            "no": new_no, "part_no": "", "part_name": "", "comment": "", "possible": "가능",
            "qty": 1, "material": "", "size": "",
            "m_5axis": 0.0, "m_4axis": 0.0, "m_3axis": 0.0, "m_lathe": 0.0,
            "m_general": 0.0, "m_finish": 0.0, "m_cmm": 0.0, "m_grind": 0.0,
            "m_jig": 0.0, "m_prog": 0.0
        })
        self.refresh_table()
        self.open_popup(new_no)

    def load_from_excel(self, show_message=True):
        file_path = self.excel_file
        if not os.path.exists(file_path):
            file_path = "견적용_입력완료.xlsx"
            if not os.path.exists(file_path):
                if show_message:
                    messagebox.showwarning("파일 미존재", "'견적용.xlsx' 파일이 존재하지 않습니다.\n새로운 기본 양식을 시작합니다.")
                return False

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_name = "기계" if "기계" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]

            loaded_data = []
            row = 7
            while True:
                part_no = ws.cell(row=row, column=2).value
                part_name = ws.cell(row=row, column=3).value

                # 10개 행 이상 읽은 뒤 데이터가 더 이상 없으면 종료
                if part_no is None and part_name is None and row >= 16:
                    break

                item = {
                    "no": len(loaded_data) + 1,
                    "part_no": str(part_no).strip() if part_no is not None else "",
                    "part_name": str(part_name).strip() if part_name is not None else "",
                    "comment": str(ws.cell(row=row, column=4).value or "").strip(),
                    "possible": str(ws.cell(row=row, column=5).value or "가능").strip(),
                    "qty": int(ws.cell(row=row, column=6).value or 1),
                    "material": str(ws.cell(row=row, column=7).value or "").strip(),
                    "size": str(ws.cell(row=row, column=8).value or "").strip(),
                    "m_5axis": float(ws.cell(row=row, column=11).value or 0.0),
                    "m_4axis": float(ws.cell(row=row, column=12).value or 0.0),
                    "m_3axis": float(ws.cell(row=row, column=13).value or 0.0),
                    "m_lathe": float(ws.cell(row=row, column=14).value or 0.0),
                    "m_general": float(ws.cell(row=row, column=15).value or 0.0),
                    "m_finish": float(ws.cell(row=row, column=16).value or 0.0),
                    "m_cmm": float(ws.cell(row=row, column=17).value or 0.0),
                    "m_grind": float(ws.cell(row=row, column=18).value or 0.0),
                    "m_jig": float(ws.cell(row=row, column=19).value or 0.0),
                    "m_prog": float(ws.cell(row=row, column=20).value or 0.0),
                }
                loaded_data.append(item)
                row += 1

            if loaded_data:
                self.data = loaded_data
                self.refresh_table()
                if show_message:
                    messagebox.showinfo("불러오기 완료", f"'{file_path}'에서 총 {len(loaded_data)}개 항목을 불러왔습니다!")
                return True
        except Exception as e:
            if show_message:
                messagebox.showerror("불러오기 오류", f"엑셀 파일을 읽는 중 오류가 발생했습니다:\n{e}")
            return False
        return False

    def open_popup(self, no):
        item_data = next((i for i in self.data if i["no"] == no), None)
        if not item_data:
            return

        popup = tk.Toplevel(self.root)
        popup.title(f"📋 [NO. {no}] 기계 시트 항목 입력")
        popup.geometry("650x670")
        popup.grab_set()

        main_frame = ttk.Frame(popup, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. 기본 정보 섹션
        lbl_info = ttk.LabelFrame(main_frame, text=" 📌 기본 정보 ", padding=10)
        lbl_info.pack(fill=tk.X, pady=5)

        fields = {}

        ttk.Label(lbl_info, text="품번 *").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        fields["part_no"] = tk.StringVar(value=item_data["part_no"])
        ttk.Entry(lbl_info, textvariable=fields["part_no"], width=18).grid(row=0, column=1, padx=5, pady=3)

        ttk.Label(lbl_info, text="품명 *").grid(row=0, column=2, sticky="e", padx=5, pady=3)
        fields["part_name"] = tk.StringVar(value=item_data["part_name"])
        ttk.Entry(lbl_info, textvariable=fields["part_name"], width=18).grid(row=0, column=3, padx=5, pady=3)

        ttk.Label(lbl_info, text="Material").grid(row=1, column=0, sticky="e", padx=5, pady=3)
        fields["material"] = tk.StringVar(value=item_data["material"])
        ttk.Entry(lbl_info, textvariable=fields["material"], width=18).grid(row=1, column=1, padx=5, pady=3)

        ttk.Label(lbl_info, text="수량(Qty)").grid(row=1, column=2, sticky="e", padx=5, pady=3)
        fields["qty"] = tk.StringVar(value=str(item_data["qty"]))
        ttk.Entry(lbl_info, textvariable=fields["qty"], width=18).grid(row=1, column=3, padx=5, pady=3)

        ttk.Label(lbl_info, text="가능여부").grid(row=2, column=0, sticky="e", padx=5, pady=3)
        fields["possible"] = tk.StringVar(value=item_data["possible"])
        ttk.Combobox(lbl_info, textvariable=fields["possible"], values=["가능", "불가", "검토필요"], width=15, state="readonly").grid(row=2, column=1, padx=5, pady=3)

        ttk.Label(lbl_info, text="Comment").grid(row=2, column=2, sticky="e", padx=5, pady=3)
        fields["comment"] = tk.StringVar(value=item_data["comment"])
        ttk.Entry(lbl_info, textvariable=fields["comment"], width=18).grid(row=2, column=3, padx=5, pady=3)

        # 2. Size (소재 규격) 입력 섹션
        lbl_size = ttk.LabelFrame(main_frame, text=" 📐 Size (소재 규격) 입력 ", padding=10)
        lbl_size.pack(fill=tk.X, pady=8)

        shape_var = tk.StringVar(value="블록")
        t_var, w_var, l_var = tk.StringVar(), tk.StringVar(), tk.StringVar()
        d_var, custom_size_var = tk.StringVar(), tk.StringVar()
        final_size_var = tk.StringVar(value=item_data["size"])

        # 기존 저장된 Size 파싱 시도
        curr_size = item_data["size"].strip()
        if "Ø" in curr_size or "D*" in curr_size:
            shape_var.set("로드")
            m = re.findall(r"[\d\.]+", curr_size)
            if len(m) >= 2:
                d_var.set(m[0])
                l_var.set(m[1])
        elif "T" in curr_size or "W" in curr_size or "*" in curr_size:
            shape_var.set("블록")
            m = re.findall(r"[\d\.]+", curr_size)
            if len(m) >= 3:
                t_var.set(m[0]); w_var.set(m[1]); l_var.set(m[2])
        elif curr_size:
            shape_var.set("직접입력")
            custom_size_var.set(curr_size)

        radio_frame = ttk.Frame(lbl_size)
        radio_frame.pack(fill=tk.X, pady=2)

        ttk.Label(radio_frame, text="형상 선택: ").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(radio_frame, text="블록 (T × W × L)", value="블록", variable=shape_var).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(radio_frame, text="로드/원봉 (D × L)", value="로드", variable=shape_var).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(radio_frame, text="직접 입력", value="직접입력", variable=shape_var).pack(side=tk.LEFT, padx=10)

        input_subframe = ttk.Frame(lbl_size, padding=5)
        input_subframe.pack(fill=tk.X, pady=5)

        def update_size_preview(*args):
            s_type = shape_var.get()
            if s_type == "블록":
                t_val, w_val, l_val = t_var.get().strip(), w_var.get().strip(), l_var.get().strip()
                if t_val or w_val or l_val:
                    final_size_var.set(f"{t_val}T * {w_val}W * {l_val}L" if t_val and w_val and l_val else f"{t_val} * {w_val} * {l_val}")
                else:
                    final_size_var.set("")
            elif s_type == "로드":
                d_val, l_val = d_var.get().strip(), l_var.get().strip()
                if d_val or l_val:
                    final_size_var.set(f"Ø{d_val} * {l_val}L" if d_val and l_val else f"Ø{d_val} * {l_val}")
                else:
                    final_size_var.set("")
            else:
                final_size_var.set(custom_size_var.get().strip())

        for v in [t_var, w_var, l_var, d_var, custom_size_var]:
            v.trace_add("write", update_size_preview)

        def render_size_inputs():
            for widget in input_subframe.winfo_children():
                widget.destroy()

            s_type = shape_var.get()
            if s_type == "블록":
                ttk.Label(input_subframe, text="T (두께):").grid(row=0, column=0, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=t_var, width=10).grid(row=0, column=1, padx=5, pady=2)
                ttk.Label(input_subframe, text="W (폭):").grid(row=0, column=2, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=w_var, width=10).grid(row=0, column=3, padx=5, pady=2)
                ttk.Label(input_subframe, text="L (길이):").grid(row=0, column=4, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=l_var, width=10).grid(row=0, column=5, padx=5, pady=2)
            elif s_type == "로드":
                ttk.Label(input_subframe, text="D (외경/Ø):").grid(row=0, column=0, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=d_var, width=12).grid(row=0, column=1, padx=5, pady=2)
                ttk.Label(input_subframe, text="L (길이):").grid(row=0, column=2, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=l_var, width=12).grid(row=0, column=3, padx=5, pady=2)
            else:
                ttk.Label(input_subframe, text="Size 규격:").grid(row=0, column=0, padx=3, pady=2)
                ttk.Entry(input_subframe, textvariable=custom_size_var, width=35).grid(row=0, column=1, padx=5, pady=2)

            update_size_preview()

        shape_var.trace_add("write", lambda *a: render_size_inputs())
        render_size_inputs()

        preview_frame = ttk.Frame(lbl_size)
        preview_frame.pack(fill=tk.X, pady=2)
        ttk.Label(preview_frame, text="최종 반영 Size: ", font=("맑은 고딕", 9, "bold")).pack(side=tk.LEFT, padx=5)
        ttk.Label(preview_frame, textvariable=final_size_var, font=("맑은 고딕", 9.5, "bold"), foreground="blue").pack(side=tk.LEFT, padx=5)

        # 3. 공수 시간 입력 섹션
        lbl_mach = ttk.LabelFrame(main_frame, text=" ⚙️ 설비 및 작업별 공수 시간 (시간 단위) ", padding=10)
        lbl_mach.pack(fill=tk.X, pady=8)

        for idx, (m_key, label_text) in enumerate(self.mach_keys):
            r = idx // 2
            c = (idx % 2) * 2
            ttk.Label(lbl_mach, text=label_text).grid(row=r, column=c, sticky="e", padx=5, pady=3)
            val_str = str(item_data[m_key]) if item_data[m_key] > 0 else ""
            var = tk.StringVar(value=val_str)
            ttk.Entry(lbl_mach, textvariable=var, width=15).grid(row=r, column=c+1, padx=5, pady=3)
            fields[m_key] = var

        def save_and_close(next_item=False):
            if not fields["part_no"].get().strip() or not fields["part_name"].get().strip():
                messagebox.showerror("입력 오류", "품번과 품명은 필수 입력 항목입니다.", parent=popup)
                return

            try:
                qty_val = int(fields["qty"].get().strip())
                if qty_val <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("입력 오류", "수량(Qty)은 1 이상의 정수로 입력하셔야 합니다.", parent=popup)
                return

            mach_vals = {}
            for m_key, _ in self.mach_keys:
                val = fields[m_key].get().strip()
                try:
                    mach_vals[m_key] = float(val) if val else 0.0
                except ValueError:
                    messagebox.showerror("입력 오류", "공수 시간은 숫자로 입력해 주세요 (예: 1.5).", parent=popup)
                    return

            item_data["part_no"] = fields["part_no"].get().strip()
            item_data["part_name"] = fields["part_name"].get().strip()
            item_data["material"] = fields["material"].get().strip()
            item_data["size"] = final_size_var.get().strip()
            item_data["comment"] = fields["comment"].get().strip()
            item_data["possible"] = fields["possible"].get().strip()
            item_data["qty"] = qty_val

            for m_key in mach_vals:
                item_data[m_key] = mach_vals[m_key]

            self.refresh_table()
            popup.destroy()

            if next_item:
                next_no = no + 1
                if not any(i["no"] == next_no for i in self.data):
                    self.data.append({
                        "no": next_no, "part_no": "", "part_name": "", "comment": "", "possible": "가능",
                        "qty": 1, "material": "", "size": "",
                        "m_5axis": 0.0, "m_4axis": 0.0, "m_3axis": 0.0, "m_lathe": 0.0,
                        "m_general": 0.0, "m_finish": 0.0, "m_cmm": 0.0, "m_grind": 0.0,
                        "m_jig": 0.0, "m_prog": 0.0
                    })
                    self.refresh_table()
                self.open_popup(next_no)

        btn_box = ttk.Frame(main_frame)
        btn_box.pack(fill=tk.X, pady=5)

        ttk.Button(btn_box, text="저장 후 다음 항목 입력 ➔", command=lambda: save_and_close(True)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_box, text="저장", command=lambda: save_and_close(False)).pack(side=tk.RIGHT, padx=5)

    def save_to_excel(self):
        file_path = self.excel_file

        # 파일 존재 시 로드, 없을 시 신규 생성
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            sheet_name = "기계" if "기계" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "기계"
            headers = ["NO", "품번", "품명", "Coment", "가능여부", "Qty", "Material", "Size",
                       "SUM(시간)", "단가(원)", "5축 NC", "4축", "3축 NC", "NC 선반", "범용", "사상", "CMM", "연삭/와이어", "치구", "프로그램"]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=6, column=col_idx, value=header)

        mach_cols = [11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
        mach_keys = [k[0] for k in self.mach_keys]

        for idx, item in enumerate(self.data):
            row = 7 + idx
            sum_time, unit_price, final_price = self.calc_row(item)

            ws.cell(row=row, column=1, value=item["no"])
            ws.cell(row=row, column=2, value=item["part_no"] if item["part_no"] else None)
            ws.cell(row=row, column=3, value=item["part_name"] if item["part_name"] else None)
            ws.cell(row=row, column=4, value=item["comment"] if item["comment"] else None)
            ws.cell(row=row, column=5, value=item["possible"])
            ws.cell(row=row, column=6, value=item["qty"])
            ws.cell(row=row, column=7, value=item["material"] if item["material"] else None)
            ws.cell(row=row, column=8, value=item["size"] if item["size"] else None)
            ws.cell(row=row, column=9, value=sum_time if (item["part_no"] or item["part_name"]) else None)
            ws.cell(row=row, column=10, value=final_price if (item["part_no"] or item["part_name"]) else None)

            for col_idx, m_key in zip(mach_cols, mach_keys):
                val = item[m_key]
                ws.cell(row=row, column=col_idx, value=val if val > 0 else None)

        output_path = "견적용_입력완료.xlsx"
        try:
            wb.save(output_path)
            wb.save(file_path)
            messagebox.showinfo("저장 완료", f"'{file_path}' 및 '{output_path}' 파일로 성공적으로 저장되었습니다!")
        except Exception as e:
            messagebox.showerror("저장 오류", f"엑셀 저장 중 오류가 발생했습니다:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = MachineEstimateApp(root)
    root.mainloop()